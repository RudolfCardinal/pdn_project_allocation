#!/usr/bin/env python

"""
pdn_project_allocation/helperfunc.py

===============================================================================

    Copyright (C) 2019-2021 Rudolf Cardinal (rudolf@pobox.com).

    This file is part of pdn_project_allocation.

    This is free software: you can redistribute it and/or modify
    it under the terms of the GNU General Public License as published by
    the Free Software Foundation, either version 3 of the License, or
    (at your option) any later version.

    This software is distributed in the hope that it will be useful,
    but WITHOUT ANY WARRANTY; without even the implied warranty of
    MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE. See the
    GNU General Public License for more details.

    You should have received a copy of the GNU General Public License
    along with this software. If not, see <https://www.gnu.org/licenses/>.

===============================================================================

Helper functions.

"""

import csv
import logging
from typing import Any, List, Sequence

from openpyxl.cell import Cell
from openpyxl.utils import get_column_letter
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from mip import Model
from mip.exceptions import SolutionNotAvailable

log = logging.getLogger(__name__)


# =============================================================================
# Helper functions
# =============================================================================

def mismatch(actual: List[Any], expected: List[Any]) -> str:
    """
    Provides text to locate a mismatch between two lists.
    """
    n_actual = len(actual)
    n_intended = len(expected)
    if n_actual != n_intended:
        return (
            f"Wrong length: actual has length {n_actual}, "
            f"intended has length {n_intended}"
        )
    for i in range(n_actual):
        if actual[i] != expected[i]:
            return f"Found {actual[i]!r} where {expected[i]!r} was expected"
    return ""


def is_empty_row(row: Sequence[Cell]) -> bool:
    """
    Is this an empty spreadsheet row?
    """
    return all(cell.value is None for cell in row)


def read_until_empty_row(ws: Worksheet) -> List[List[Any]]:
    """
    Reads a spreadsheet until the first empty line.
    (Helpful because Excel spreadsheets are sometimes seen as having 1048576
    rows when they don't really).
    """
    rows = []  # type: List[List[Any]]
    for row in ws.iter_rows():
        if is_empty_row(row):
            break
        rows.append([cell.value for cell in row])
    return rows


def report_on_model(m: Model,
                    loglevel: int = logging.WARNING,
                    solution_only: bool = False) -> None:
    """
    Shows detail of a MIP model to the log.
    """
    lines = ["Model:", "", "- Variables:", ""]
    try:
        for v in m.vars:
            lines.append(f"{v.title} == {v.x}")
    except SolutionNotAvailable:
        if solution_only:
            raise
        for v in m.vars:
            lines.append(f"{v.title}")
    if not solution_only:
        lines += ["", "- Objective:", ""]
        lines.append(str(m.objective.sense))
        lines.append(str(m.objective))
        lines += ["", "- Constraints:", ""]
        for c in m.constrs:
            lines.append(str(c))
    log.log(loglevel, "\n".join(lines))


def csv_to_supervisor_names(csv_names: str) -> List[str]:
    """
    From a string of comma-separated supervisor names like ``Dr Smith, Dr
    Jones``, return the supervisor names, like ``['Dr Smith', 'Dr Jones']``.
    """
    if not csv:
        return []
    return [x.strip() for x in csv_names.split(",")]


def supervisor_names_to_csv(names: List[str]) -> str:
    """
    Opposite of :func:`csv_to_supervisor_names`.
    """
    return ", ".join(names)


def autosize_openpyxl_column(ws: Worksheet, col_number: int) -> None:
    """
    Automatically resize a single column to its contents. See below.
    """
    col_width = 0
    for row in ws.rows:
        cell = row[col_number]
        if cell.value:
            text = str(cell.value)
            text_width = len(text)
            col_width = max(col_width, text_width)
    ws.column_dimensions[get_column_letter(col_number + 1)].width = col_width


def autosize_openpyxl_worksheet_columns(ws: Worksheet) -> None:
    """
    Automatically resize column sizes to their contents. See
    
    - https://stackoverflow.com/questions/13197574/openpyxl-adjust-column-width-size
    - https://stackoverflow.com/questions/60248319/how-to-set-column-width-to-bestfit-in-openpyxl
    """  # noqa

    # Method 1 -- OK but overestimates size.
    # Better would be to ask for actual size with current font.
    dims = {}
    for row in ws.rows:
        for cell in row:
            if cell.value:
                text = str(cell.value)
                text_width = len(text)  # the poor approximation
                dims[cell.column_letter] = max(
                    dims.get(cell.column_letter, 0), text_width)
    for col, value in dims.items():
        ws.column_dimensions[col].width = value

    # Method 2 -- doesn't work
    # column_letters = tuple(
    #     get_column_letter(col_number + 1)
    #     for col_number in range(ws.max_column)
    # )
    # for column_letter in column_letters:
    #     # noinspection PyPep8Naming
    #     ws.column_dimensions[column_letter].bestFit = True


def autosize_openpyxl_columns_all_sheets(wb: Workbook) -> None:
    """
    Autosize columns for all sheets in a workbook.
    """
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        autosize_openpyxl_worksheet_columns(ws)
