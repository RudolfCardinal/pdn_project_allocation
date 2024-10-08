#!/usr/bin/env python

"""
pdn_project_allocation/problem.py

===============================================================================

    Copyright (C) 2019 Rudolf Cardinal (rudolf@pobox.com).

    This file is part of pdn_project_allocation.

    This is free software: you can redistribute it and/or modify
    it under the terms of the GNU General Public License as published by
    the Free Software Foundation, either version 3 of the License, or
    (at your option) any later version.

    This software is distributed in the hope that it will be useful,
    but WITHOUT ANY WARRANTY; without even the implied warranty of
    MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE. See the
    GNU General Public License for more details.

    You should have received a copy of the GNU General Public License
    along with this software. If not, see <https://www.gnu.org/licenses/>.

===============================================================================

Problem class, which also solves the problem.

"""

from collections import OrderedDict
from itertools import product
import logging
import os
import random
from typing import (
    Any,
    Dict,
    Generator,
    List,
    Optional,
    Sequence,
    Set,
    Tuple,
    Union,
)

from openpyxl.reader.excel import load_workbook
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from matching.games.student_allocation import (
    Project as MGProject,
    Student as MGStudent,
    StudentAllocation as MGStudentAllocation,
    Supervisor as MGSupervisor,
)
from mip import BINARY, LinExpr, minimize, Model, Var, xsum

from pdn_project_allocation.config import Config
from pdn_project_allocation.constants import (
    EXT_XLSX,
    FALSE_VALUES,
    MISSING_VALUES,
    OptimizeMethod,
    SheetHeadings,
    SheetNames,
    Switches,
    TRUE_VALUES,
)
from pdn_project_allocation.eligibility import Eligibility
from pdn_project_allocation.helperfunc import (
    autosize_openpyxl_column,
    autosize_openpyxl_worksheet_columns,
    bold_first_row,
    csv_to_supervisor_names,
    mismatch,
    read_until_empty_row,
    report_on_model,
)
from pdn_project_allocation.project import Project
from pdn_project_allocation.solution import Solution
from pdn_project_allocation.student import Student
from pdn_project_allocation.supervisor import Supervisor

log = logging.getLogger(__name__)


# =============================================================================
# Constants
# =============================================================================

ALMOST_ONE = 0.99


# =============================================================================
# Problem
# =============================================================================


class Problem:
    """
    Represents the problem (and solves it) -- projects (with their supervisor's
    preferences for students), students (with their preferences for projects),
    and eligibility (which students are allowed to do which project?).
    """

    def __init__(
        self,
        supervisors: List[Supervisor],
        projects: List[Project],
        students: List[Student],
        config: Config,
        eligibility: Eligibility = None,
    ) -> None:
        """
        Args:
            supervisors:
                List of project supervisors.
            projects:
                List of projects (with supervisor preference information).
            students:
                List of students (with their project preferences).
            config:
                Master config object
            eligibility:
                Dictionary of the form eligibility[student][project] = bool.

        Note that the students and projects are put into a "deterministic
        random" order, i.e. deterministically sorted, then shuffled (but with a
        globally fixed random number generator seed). That ensures fairness and
        consistency.
        """
        self.supervisors = supervisors
        self.projects = projects
        self.students = students
        self.config = config
        self.eligibility = eligibility or Eligibility(students, projects)
        self.eligibility.assert_valid()
        # Fix the order:
        for i, p in enumerate(self.projects):
            p.original_order = i
        for i, s in enumerate(self.students):
            s.original_order = i
        if not config.no_shuffle:
            self.students.sort()
            random.shuffle(self.students)
            self.projects.sort()
            random.shuffle(self.projects)

    # -------------------------------------------------------------------------
    # Representations
    # -------------------------------------------------------------------------

    def __str__(self) -> str:
        """
        We re-sort the output for display purposes.
        """
        supervisors = "\n".join(
            sv.description() for sv in self.sorted_supervisors()
        )
        projects = "\n".join(p.description() for p in self.sorted_projects())
        students = "\n".join(s.description() for s in self.sorted_students())
        return (
            f"Problem:\n"
            f"\n"
            f"- Supervisors:\n\n{supervisors}\n"
            f"\n"
            f"- Projects:\n\n{projects}\n"
            f"\n"
            f"- Students:\n\n{students}\n"
            f"\n"
            f"- Eligibility:\n\n{self.eligibility}\n"
        )

    # -------------------------------------------------------------------------
    # Information
    # -------------------------------------------------------------------------

    def sorted_supervisors(self) -> List[Supervisor]:
        """
        Supervisors, sorted by number.
        """
        return sorted(self.supervisors, key=lambda sv: sv.number)

    def sorted_students(self) -> List[Student]:
        """
        Students, sorted by number.
        """
        return sorted(self.students, key=lambda s: s.number)

    def sorted_projects(self) -> List[Project]:
        """
        Projects, sorted by number.
        """
        return sorted(self.projects, key=lambda p: p.number)

    def n_supervisors(self) -> int:
        """
        Number of supervisors.
        """
        return len(self.supervisors)

    def n_students(self) -> int:
        """
        Number of students.
        """
        return len(self.students)

    def n_projects(self) -> int:
        """
        Number of projects.
        """
        return len(self.projects)

    def students_who_chose(self, project: Project) -> List[Student]:
        """
        All students who ranked this project in some way.
        """
        return [
            s
            for s in self.students
            if s.preferences.actively_expressed_preference_for(project)
        ]

    def gen_student_project_pairs_where_student_chose_project(
        self,
    ) -> Generator[Tuple[Student, Project], None, None]:
        """
        Generate ``student, project`` tuples where the student expressed some
        interest in the project.
        """
        for s in self.students:
            for p in s.preferences.items_explicitly_ranked():
                yield s, p

    @staticmethod
    def is_student_interested(student: Student, project: Project) -> bool:
        """
        UNUSED. Is the student interested in this project?
        """
        return student.preferences.actively_expressed_preference_for(project)

    def are_preferences_strict_over_relevant_combos(self) -> bool:
        """
        Are all preferences strict, across combinations that matter?
        """
        # Students should strictly order their projects:
        for s in self.students:
            if not s.preferences.is_strict_over_expressed_preferences():
                return False
        # Supervisors should strictly order the students who expressed an
        # interest in their projects:
        for p in self.projects:
            students = self.students_who_chose(p)
            if not p.supervisor_preferences.is_strict_over(students):
                return False
        return True

    def gen_better_projects(
        self, student: Student, project: Project
    ) -> Generator[Project, None, None]:
        """
        Generates projects that this student prefers over the specified one
        (and for which they're eligible).
        """
        current_dissatisfaction = student.dissatisfaction(project)
        for p in self.projects:
            if not self.eligibility.is_eligible(student, p):
                continue
            new_dissatisfaction = student.dissatisfaction(p)
            if new_dissatisfaction < current_dissatisfaction:
                yield p

    def gen_better_students(
        self, project: Project, student: Student
    ) -> Generator[Student, None, None]:
        """
        Generates students that this project prefers over the specified one
        (and for which they're eligible).
        """
        current_dissatisfaction = project.dissatisfaction(student)
        for s in self.students:
            if not self.eligibility.is_eligible(s, project):
                continue
            new_dissatisfaction = project.dissatisfaction(s)
            if new_dissatisfaction < current_dissatisfaction:
                yield s

    def gen_worse_students(
        self, project: Project, student: Student
    ) -> Generator[Student, None, None]:
        """
        Generates students that this project prefers LESS THAN the specified
        one (and for which they're eligible).
        """
        current_dissatisfaction = project.dissatisfaction(student)
        for s in self.students:
            if not self.eligibility.is_eligible(s, project):
                continue
            new_dissatisfaction = project.dissatisfaction(s)
            if new_dissatisfaction > current_dissatisfaction:
                yield s

    # -------------------------------------------------------------------------
    # Read data
    # -------------------------------------------------------------------------

    @classmethod
    def read_data(cls, config: Config) -> "Problem":
        """
        Reads a file, autodetecting its format, and returning the
        :class:`Problem`.
        """
        # File type?
        _, ext = os.path.splitext(config.filename)
        if ext == EXT_XLSX:
            return cls.read_data_xlsx(config)
        else:
            raise ValueError(
                f"Don't know how to read file type {ext!r} "
                f"for {config.filename!r}"
            )

    # noinspection DuplicatedCode
    @classmethod
    def read_data_xlsx(cls, config: Config) -> "Problem":
        """
        Reads a :class:`Problem` from an Excel XLSX file.
        """

        log.info(f"Reading XLSX file: {config.filename}")
        wb = load_workbook(
            config.filename,
            read_only=True,
            keep_vba=False,
            data_only=True,
            keep_links=False,
        )
        supervisors, sv_name_to_supervisor = cls._read_supervisors(wb)
        projects = cls._read_projects(wb, config, sv_name_to_supervisor)
        students = cls._read_students(wb, config, projects)
        cls._read_supervisor_preferences(wb, config, projects, students)
        eligibility = cls._read_eligibility(wb, config, projects, students)
        wb.close()
        log.info("... finished reading")

        # Create and return the Problem object
        return Problem(
            supervisors=supervisors,
            projects=projects,
            students=students,
            eligibility=eligibility,
            config=config,
        )

    @classmethod
    def _assert(
        cls,
        assertion: Any,
        error_msg: str,
        errors_so_far: List[str] = None,
        proceed_if_error: bool = False,
    ) -> bool:
        """
        Check an assertion, return "ok?" as a Boolean, or raise ValueError.
        Provides the facility to collect several errors, for user helpfulness
        ("fix these five problems", rather than "here's the first problem, come
        back when it's fixed; here's the next problem...").
        """
        if bool(assertion):
            return True
        if errors_so_far is not None:
            errors_so_far.append(error_msg)
        if proceed_if_error:
            assert (
                errors_so_far is not None
            ), "Bug: must provide errors_so_far if using proceed_if_error"
            return False
        else:
            exception_text = (
                "\n".join(errors_so_far)
                if errors_so_far is not None
                else error_msg
            )
            log.critical(exception_text)
            raise ValueError(exception_text)

    @classmethod
    def _complete_assertions(cls, errors_so_far, context: str = "") -> None:
        """
        If there are errors, raise ValueError.
        """
        if errors_so_far:
            msg = (
                f"There were spreadsheet errors, in this context: {context}\n"
                + "\n".join("- " + x for x in errors_so_far)
            )
            log.critical(msg)
            raise ValueError(msg)

    @classmethod
    def _read_supervisors(
        cls, wb: Workbook
    ) -> Tuple[List[Supervisor], Dict[str, Supervisor]]:
        """
        Read supervisors from a spreadsheet.
        """
        log.info("... reading supervisors...")
        supervisors = []  # type: List[Supervisor]
        sv_name_to_supervisor = {}  # type: Dict[str, Supervisor]
        # This will raise an error if the named sheet does not exist:
        ws_supervisors = wb[SheetNames.SUPERVISORS]  # type: Worksheet
        expected_headings = [
            SheetHeadings.SUPERVISOR,
            SheetHeadings.MAX_NUMBER_OF_PROJECTS,
            SheetHeadings.MAX_NUMBER_OF_STUDENTS,
        ]
        obtained_headings = [c.value for c in ws_supervisors["A1:C1"][0]]
        cls._assert(
            obtained_headings == expected_headings,
            (
                f"Bad headings to worksheet {SheetNames.SUPERVISORS!r}; "
                f"expected {expected_headings!r}, got {obtained_headings!r}"
            ),
        )
        errors_so_far = []  # type: List[str]
        sv_rows = read_until_empty_row(ws_supervisors)
        for row_number, row in enumerate(sv_rows[1:], start=2):
            supervisor_number = row_number - 1
            supervisor_name = row[0].strip()
            ok = cls._assert(
                supervisor_name,
                (
                    f"Missing supervisor name in {SheetNames.SUPERVISORS!r} "
                    f"row {row_number}"
                ),
                errors_so_far=errors_so_far,
                proceed_if_error=True,
            )
            ok = (
                cls._assert(
                    supervisor_name not in sv_name_to_supervisor,
                    (
                        f"Duplicate supervisor name in "
                        f"{SheetNames.SUPERVISORS!r} "
                        f"row {row_number}: {supervisor_name!r}"
                    ),
                    errors_so_far=errors_so_far,
                    proceed_if_error=True,
                )
                and ok
            )
            max_n_projects = row[1]
            ok = (
                cls._assert(
                    max_n_projects is None or isinstance(max_n_projects, int),
                    (
                        f"Bad max_n_projects in {SheetNames.SUPERVISORS!r} "
                        f"row {row_number}; is {max_n_projects!r}"
                    ),
                    errors_so_far=errors_so_far,
                    proceed_if_error=True,
                )
                and ok
            )
            max_n_students = row[2]
            ok = (
                cls._assert(
                    max_n_students is None or isinstance(max_n_students, int),
                    (
                        f"Bad max_n_students in {SheetNames.SUPERVISORS!r} "
                        f"row {row_number}; is {max_n_students!r}"
                    ),
                    errors_so_far=errors_so_far,
                    proceed_if_error=True,
                )
                and ok
            )
            if not ok:
                continue
            new_supervisor = Supervisor(
                name=supervisor_name,
                number=supervisor_number,
                max_n_projects=max_n_projects,
                max_n_students=max_n_students,
            )
            sv_name_to_supervisor[supervisor_name] = new_supervisor
            supervisors.append(new_supervisor)
        n_supervisors = len(supervisors)
        cls._assert(
            n_supervisors,
            "No supervisors defined!",
            errors_so_far=errors_so_far,
        )
        cls._complete_assertions(errors_so_far, context=SheetNames.SUPERVISORS)
        log.info(f"Number of supervisors: {n_supervisors}")
        return supervisors, sv_name_to_supervisor

    @classmethod
    def _read_projects(
        cls,
        wb: Workbook,
        config: Config,
        sv_name_to_supervisor: Dict[str, Supervisor],
    ) -> List[Project]:
        """
        Read projects from a spreadsheet.
        """
        log.info("... reading projects...")
        projects = []  # type: List[Project]
        ws_projects = wb[SheetNames.PROJECTS]  # type: Worksheet
        project_names = []  # type: List[str]
        expected_headings = [
            SheetHeadings.PROJECT,
            SheetHeadings.MAX_NUMBER_OF_STUDENTS,
            SheetHeadings.SUPERVISOR,
        ]
        obtained_headings = [c.value for c in ws_projects["A1:C1"][0]]
        cls._assert(
            obtained_headings == expected_headings,
            (
                f"Bad headings to worksheet {SheetNames.PROJECTS!r}; expected "
                f"{expected_headings!r}, got {obtained_headings!r}"
            ),
        )
        # log.debug(f"Projects: max_row = {ws_projects.max_row}")
        errors_so_far = []  # type: List[str]
        p_rows = read_until_empty_row(ws_projects)
        for row_number, row in enumerate(p_rows[1:], start=2):
            project_number = row_number - 1
            project_name = row[0].strip()
            ok = cls._assert(
                project_name,
                (
                    f"Missing project name in {SheetNames.PROJECTS!r} "
                    f"row {row_number}"
                ),
                errors_so_far=errors_so_far,
                proceed_if_error=True,
            )
            ok = (
                cls._assert(
                    project_name not in project_names,
                    (
                        f"Duplicate project name in {SheetNames.PROJECTS!r} "
                        f"row {row_number}: {project_name!r}"
                    ),
                    errors_so_far=errors_so_far,
                    proceed_if_error=True,
                )
                and ok
            )
            max_n_students = row[1]
            ok = (
                cls._assert(
                    isinstance(max_n_students, int),
                    (
                        f"Bad max_n_students in {SheetNames.PROJECTS!r} "
                        f"row {row_number}; is {max_n_students!r}"
                    ),
                    errors_so_far=errors_so_far,
                    proceed_if_error=True,
                )
                and ok
            )
            supervisor_names = csv_to_supervisor_names(row[2])
            this_project_supervisors = []
            for sv_name in supervisor_names:
                ok = (
                    cls._assert(
                        sv_name in sv_name_to_supervisor,
                        (
                            f"Unknown supervisor in {SheetNames.PROJECTS!r} "
                            f"row {row_number}: {sv_name!r} "
                            f"(full cell is {row[2]!r})"
                        ),
                        errors_so_far=errors_so_far,
                        proceed_if_error=True,
                    )
                    and ok
                )
                this_project_supervisors.append(sv_name_to_supervisor[sv_name])
            project_names.append(project_name)
            if not ok:
                continue
            projects.append(
                Project(
                    title=project_name,
                    number=project_number,
                    supervisors=this_project_supervisors,
                    max_n_students=max_n_students,
                    allow_defunct_projects=config.allow_defunct_projects,
                )
            )
        n_projects = len(projects)
        cls._assert(
            n_projects, "No projects defined!", errors_so_far=errors_so_far
        )
        cls._complete_assertions(errors_so_far, context=SheetNames.PROJECTS)
        log.info(f"Number of projects: {n_projects}")
        return projects

    @classmethod
    def get_pref_val(cls, x_: Any) -> Union[int, float, None]:
        """
        Raises ValueError if it's not a preference quantity.
        """
        if x_ is None:
            return None
        if isinstance(x_, str) and not x_.strip():
            # Empty string or string containing spaces.
            return None
        if isinstance(x_, (int, float)):
            return x_
        raise ValueError(f"Bad preference: {x_!r}")

    @classmethod
    def _read_students(
        cls, wb: Workbook, config: Config, projects: List[Project]
    ) -> List[Student]:
        """
        Read students and their preferences from a spreadsheet.
        """
        log.info("... reading students and their preferences...")
        students = []  # type: List[Student]
        student_names = []  # type: List[str]
        ws_students = wb[
            SheetNames.STUDENT_PREFERENCES
        ]  # type: Worksheet  # noqa
        stp_rows = read_until_empty_row(ws_students)
        # Check project headings
        n_projects = len(projects)
        for i in range(n_projects):
            cls._assert(
                stp_rows[0][i + 1].strip() == projects[i].title,
                (
                    f"First row of {SheetNames.STUDENT_PREFERENCES!r} sheet "
                    f"must contain all project names in the same order as in "
                    f"the {SheetNames.PROJECTS!r} sheet. For project {i + 1}, "
                    f"the project name is {projects[i].title!r}, but the "
                    f"column heading is {stp_rows[0][i + 1]!r}."
                ),
            )
        expected_row_len = n_projects + 1
        errors_so_far = []  # type: List[str]
        for row_number, row in enumerate(stp_rows[1:], start=2):
            student_number = row_number - 1
            student_name = row[0].strip()
            # 2023-09-25: often there are blank rows at the end. That's not
            # unreasonable, so truncate.
            row = row[:expected_row_len]
            # Any residual error here is therefore a row that's too short.
            ok = cls._assert(
                len(row) == expected_row_len,
                (
                    f"In {SheetNames.STUDENT_PREFERENCES!r}, student on row "
                    f"{row_number} (named {student_name!r}) has a preference "
                    f"row of the wrong length (expected {expected_row_len}, "
                    f"got {len(row)})."
                ),
            )
            ok = (
                cls._assert(
                    student_name not in student_names,
                    (
                        f"Duplicate student name in "
                        f"{SheetNames.STUDENT_PREFERENCES!r} "
                        f"row {row_number}: {student_name!r}"
                    ),
                    errors_so_far=errors_so_far,
                    proceed_if_error=True,
                )
                and ok
            )
            student_preferences = OrderedDict()  # type: Dict[Project, int]
            for project_number, pref_contents in enumerate(row[1:], start=1):
                pref = None
                try:
                    pref = cls.get_pref_val(pref_contents)  # may raise
                    if config.allow_student_preference_ties:
                        # Float permitted here. But integers also permitted.
                        pref_ok = isinstance(pref, (type(None), int, float))
                    else:
                        # Float not OK here.
                        pref_ok = isinstance(pref, (type(None), int))
                    if not pref_ok:
                        raise ValueError
                except ValueError:
                    ok = cls._assert(
                        False,
                        (
                            f"Bad preference for student {student_name!r} in "
                            f"{SheetNames.STUDENT_PREFERENCES!r} "
                            f"row {row_number}: {pref_contents!r}"
                        ),
                        errors_so_far=errors_so_far,
                        proceed_if_error=True,
                    )
                project = projects[project_number - 1]
                if pref is not None:
                    student_preferences[project] = pref
            student_names.append(student_name)
            if not ok:
                continue
            preferred_amongst_unranked = []  # type: List[Project]
            if config.assume_supervisor_affinity:
                preferred_supervisors = set()  # type: Set[Supervisor]
                for project in student_preferences.keys():
                    for supervisor in project.supervisors:
                        preferred_supervisors.add(supervisor)
                for project in projects:
                    if project not in student_preferences.keys() and any(
                        supervisor in preferred_supervisors
                        for supervisor in project.supervisors
                    ):
                        preferred_amongst_unranked.append(project)
            try:
                new_student = Student(
                    name=student_name,
                    number=student_number,
                    preferences=student_preferences,
                    n_projects=n_projects,
                    allow_ties=config.allow_student_preference_ties,
                    preference_power=config.preference_power,
                    input_rank_notation=config.input_rank_notation,
                    preferred_amongst_unranked=preferred_amongst_unranked,
                )
                students.append(new_student)
            except Exception as exc:
                cls._assert(
                    False,
                    repr(exc),
                    errors_so_far=errors_so_far,
                    proceed_if_error=True,
                )
        n_students = len(students)
        cls._assert(
            n_students >= 1,
            "No students defined!",
            errors_so_far=errors_so_far,
        )
        cls._complete_assertions(
            errors_so_far, context=SheetNames.STUDENT_PREFERENCES
        )
        log.info(f"Number of students: {n_students}")
        return students

    @classmethod
    def _read_supervisor_preferences(
        cls,
        wb: Workbook,
        config: Config,
        projects: List[Project],
        students: List[Student],
    ) -> None:
        """
        Read supervisor-provided per-project preferences (for students) from a
        spreadsheet, writing them to the known projects.
        """
        log.info("... reading supervisor preferences...")
        ws_supervisorprefs = wb[
            SheetNames.SUPERVISOR_PREFERENCES
        ]  # type: Worksheet  # noqa
        # Accessing cells by (row, column) index is ridiculously slow here, and
        # the time is spent in the internals of openpyxl; specifically, in
        # xml.etree.ElementTree.XMLParser.feed(). That's true even after
        # install lxml as recommended, and specifying the "simple read-only"
        # options. So, it is **much** faster to load all the values like this
        # and then operate on the copies (e.g. ~6 seconds becomes ~1 ms):
        svp_rows = read_until_empty_row(ws_supervisorprefs)
        # ... index as : svp_rows[row_zero_based][column_zero_based]

        n_projects = len(projects)
        n_students = len(students)

        # Check project headings
        for i in range(n_projects):
            cls._assert(
                svp_rows[0][i + 1].strip() == projects[i].title,
                (
                    f"First row of {SheetNames.SUPERVISOR_PREFERENCES!r} "
                    f"sheet must contain all project names in the same order "
                    f"as in the {SheetNames.PROJECTS!r} sheet. For project "
                    f"{i + 1}, the project name is {projects[i].title!r}, but "
                    f"the column heading is {svp_rows[0][i + 1]!r}."
                ),
            )
        # Check student names
        cls._assert(
            len(svp_rows) == 1 + n_students,
            (
                f"Sheet {SheetNames.SUPERVISOR_PREFERENCES!r} should have "
                f"{1 + n_students} rows (one header row plus {n_students} "
                f"rows for students). Yours has {len(svp_rows)}."
            ),
        )
        # noinspection DuplicatedCode
        _sn_from_sheet = [
            svp_rows[i + 1][0].strip() for i in range(n_students)
        ]
        _sn_from_students = [students[i].name for i in range(n_students)]
        cls._assert(
            _sn_from_sheet == _sn_from_students,
            (
                f"First column of {SheetNames.SUPERVISOR_PREFERENCES!r} sheet "
                f"must contain all student names in the same order as in the "
                f"{SheetNames.STUDENT_PREFERENCES!r} sheet. Mismatch is: "
                f"{mismatch(_sn_from_sheet, _sn_from_students)}"
            ),
        )
        # Read preferences
        errors_so_far = []  # type: List[str]
        for pcol, project in enumerate(projects, start=2):
            supervisor_prefs = OrderedDict()  # type: Dict[Student, int]
            ok = True
            for srow, student in enumerate(students, start=2):
                pref_contents = svp_rows[srow - 1][pcol - 1]
                try:
                    pref = cls.get_pref_val(pref_contents)
                    supervisor_prefs[student] = pref
                except ValueError:
                    ok = cls._assert(
                        False,
                        (
                            f"Bad preference at row={srow}, col={pcol} in "
                            f"{SheetNames.SUPERVISOR_PREFERENCES!r}: "
                            f"{pref_contents!r}"
                        ),
                    )
            if ok:
                try:
                    project.set_supervisor_preferences(
                        n_students=n_students,
                        preferences=supervisor_prefs,
                        allow_ties=config.allow_supervisor_preference_ties,
                        preference_power=config.preference_power,
                        input_rank_notation=config.input_rank_notation,
                    )
                except Exception as exc:
                    cls._assert(
                        False,
                        repr(exc),
                        errors_so_far=errors_so_far,
                        proceed_if_error=True,
                    )
        cls._complete_assertions(
            errors_so_far, context=SheetNames.SUPERVISOR_PREFERENCES
        )

    @classmethod
    def _read_eligibility(
        cls,
        wb: Workbook,
        config: Config,
        projects: List[Project],
        students: List[Student],
    ) -> Eligibility:
        """
        Read student eligibility for projects from a spreadsheet.
        """
        log.info("... reading eligibility...")
        eligibility = Eligibility(
            students=students,
            projects=projects,
            default_eligibility=True,
            allow_defunct_projects=config.allow_defunct_projects,
        )
        if SheetNames.ELIGIBILITY not in wb:
            # Not provided; we just use the default.
            return eligibility

        ws_eligibility = wb[SheetNames.ELIGIBILITY]
        el_rows = read_until_empty_row(ws_eligibility)
        # ... index as : el_rows[row_zero_based][column_zero_based]
        n_projects = len(projects)
        n_students = len(students)
        # Check project headings
        for i in range(n_projects):
            cls._assert(
                el_rows[0][i + 1].strip() == projects[i].title,
                (
                    f"First row of {SheetNames.ELIGIBILITY!r} sheet must "
                    f"contain all project names in the same order as in the "
                    f"{SheetNames.PROJECTS!r} sheet. For project {i + 1}, "
                    f"project name is {projects[i].title!r}, but column "
                    f"heading is {el_rows[0][i + 1]!r}."
                ),
            )
        # Check student names
        # noinspection DuplicatedCode
        _sn_from_sheet = [el_rows[i + 1][0].strip() for i in range(n_students)]
        _sn_from_students = [
            students[i].name for i in range(n_students)
        ]  # noqa
        cls._assert(
            _sn_from_sheet == _sn_from_students,
            (
                f"First column of {SheetNames.ELIGIBILITY!r} sheet "
                f"must contain all student names in the same order as in the "
                f"{SheetNames.STUDENT_PREFERENCES!r} sheet. Mismatch is: "
                f"{mismatch(_sn_from_sheet, _sn_from_students)}"
            ),
        )
        # Read eligibility
        if config.student_must_have_choice:
            log.warning(
                f"Using option {Switches.STUDENT_MUST_HAVE_CHOICE}; "
                f"eligibility will be set to 'no' for all projects that a "
                f"student did not explicitly choose"
            )
        errors_so_far = []  # type: List[str]
        for pcol, project in enumerate(projects, start=2):
            for srow, student in enumerate(students, start=2):
                ok = True
                eligibility_val = el_rows[srow - 1][pcol - 1]
                if isinstance(eligibility_val, str):
                    eligibility_val = eligibility_val.strip()
                if eligibility_val in TRUE_VALUES:
                    eligible = True
                elif eligibility_val in FALSE_VALUES:
                    eligible = False
                elif (
                    eligibility_val in MISSING_VALUES
                    and config.missing_eligibility is not None
                ):
                    eligible = config.missing_eligibility
                else:
                    ok = cls._assert(
                        False,
                        (
                            f"Eligibility value {eligibility_val!r} (at row "
                            f"{srow}, column {pcol}) is invalid; use one of "
                            f"{TRUE_VALUES} for 'eligible', or one of "
                            f"{FALSE_VALUES} for 'ineligible'. The meaning of "
                            f"{MISSING_VALUES} is configurable; see "
                            f"{Switches.MISSING_ELIGIBILITY!r}."
                        ),
                        errors_so_far=errors_so_far,
                        proceed_if_error=True,
                    )
                    eligible = None
                    # ... won't be used; just stops PyCharm complaining
                if not ok:
                    continue
                if (
                    config.student_must_have_choice
                    and not student.explicitly_ranked_project(project)
                ):
                    eligible = False
                eligibility.set_eligibility(student, project, eligible)
        cls._complete_assertions(errors_so_far, context=SheetNames.ELIGIBILITY)
        return eligibility

    # -------------------------------------------------------------------------
    # Save data
    # -------------------------------------------------------------------------

    # noinspection DuplicatedCode
    def write_to_xlsx_workbook(
        self, wb: Workbook, with_internal_prefs: bool = False
    ) -> None:
        """
        Writes the problem data to a spreadsheet (so it can be saved alongside
        the solution).

        Args:
            wb:
                A :class:`openpyxl.workbook.workbook.Workbook` to which to
                write.
            with_internal_prefs:
                Add a sheet showing internal calculations, i.e. preferences for
                projects not explicitly chosen.
        """
        sorted_projects = self.sorted_projects()
        sorted_students = self.sorted_students()

        # ---------------------------------------------------------------------
        # Supervisors
        # ---------------------------------------------------------------------

        supervisor_sheet = wb.create_sheet(SheetNames.SUPERVISORS)
        supervisor_sheet.append(
            [
                SheetHeadings.SUPERVISOR,
                SheetHeadings.MAX_NUMBER_OF_PROJECTS,
                SheetHeadings.MAX_NUMBER_OF_STUDENTS,
            ]
        )
        for sv in self.sorted_supervisors():
            supervisor_sheet.append(
                [sv.name, sv.max_n_projects, sv.max_n_students]
            )
        autosize_openpyxl_worksheet_columns(supervisor_sheet)
        bold_first_row(supervisor_sheet)

        # ---------------------------------------------------------------------
        # Projects
        # ---------------------------------------------------------------------

        project_sheet = wb.create_sheet(SheetNames.PROJECTS)
        project_sheet.append(
            [
                SheetHeadings.PROJECT,
                SheetHeadings.MAX_NUMBER_OF_STUDENTS,
                SheetHeadings.SUPERVISOR,
            ]
        )
        for p in sorted_projects:
            project_sheet.append(
                [
                    p.title,
                    p.max_n_students,
                    p.supervisor_name(),
                ]
            )
        autosize_openpyxl_worksheet_columns(project_sheet)
        bold_first_row(project_sheet)

        # ---------------------------------------------------------------------
        # Students
        # ---------------------------------------------------------------------

        student_sheet = wb.create_sheet(SheetNames.STUDENT_PREFERENCES)
        student_sheet.append(
            [SheetHeadings.STUDENT] + [p.title for p in sorted_projects]
        )
        for s in sorted_students:
            student_sheet.append(
                [s.name]
                + [s.preferences.raw_preference(p) for p in sorted_projects]
            )
        autosize_openpyxl_column(student_sheet, 0)
        # Not all columns (lengthy project titles at top).
        bold_first_row(student_sheet)

        # ---------------------------------------------------------------------
        # Supervisor preferences
        # ---------------------------------------------------------------------

        supervisor_sheet = wb.create_sheet(SheetNames.SUPERVISOR_PREFERENCES)
        supervisor_sheet.append(
            [SheetHeadings.STUDENT] + [p.title for p in sorted_projects]
        )
        for s in sorted_students:
            # noinspection PyTypeChecker
            supervisor_sheet.append(
                [s.name]
                + [
                    p.supervisor_preferences.raw_preference(s)
                    for p in sorted_projects
                ]
            )
        autosize_openpyxl_column(supervisor_sheet, 0)
        # Not all columns (lengthy project titles at top).
        bold_first_row(supervisor_sheet)

        # ---------------------------------------------------------------------
        # Eligibility
        # ---------------------------------------------------------------------

        eligibility_sheet = wb.create_sheet(SheetNames.ELIGIBILITY)
        eligibility_sheet.append(
            [SheetHeadings.STUDENT] + [p.title for p in sorted_projects]
        )
        for s in sorted_students:
            # noinspection PyTypeChecker
            eligibility_sheet.append(
                [s.name]
                + [
                    int(self.eligibility.is_eligible(s, p))
                    for p in sorted_projects
                ]
            )
        autosize_openpyxl_column(eligibility_sheet, 0)
        # Not all columns (lengthy project titles at top).
        bold_first_row(eligibility_sheet)

        # ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
        # Internal (full) version of student preferences
        # ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
        if with_internal_prefs:
            sti = wb.create_sheet(SheetNames.STUDENT_PREFERENCES_INTERNAL)
            sti.append(
                [SheetHeadings.STUDENT] + [p.title for p in sorted_projects]
            )
            for s in sorted_students:
                # noinspection PyTypeChecker
                sti.append(
                    [s.name]
                    + [
                        s.exponentiated_dissatisfaction(p)
                        for p in sorted_projects
                    ]
                )
            autosize_openpyxl_column(sti, 0)
            bold_first_row(sti)

        # ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
        # Internal (full) version of supervisor preferences
        # ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
        if with_internal_prefs:
            spi = wb.create_sheet(SheetNames.SUPERVISOR_PREFERENCES_INTERNAL)
            spi.append(
                [SheetHeadings.STUDENT] + [p.title for p in sorted_projects]
            )
            for s in sorted_students:
                # noinspection PyTypeChecker
                spi.append(
                    [s.name]
                    + [
                        p.exponentiated_dissatisfaction(s)
                        for p in sorted_projects
                    ]
                )
            autosize_openpyxl_column(spi, 0)
            bold_first_row(spi)

    # -------------------------------------------------------------------------
    # Solver entry point
    # -------------------------------------------------------------------------

    def best_solution(self) -> Optional[Solution]:
        """
        Return the best solution.
        """
        method = self.config.optimize_method
        if method == OptimizeMethod.MINIMIZE_DISSATISFACTION:
            return self.best_solution_mip(enforce_stability=False)
        elif method == OptimizeMethod.MINIMIZE_DISSATISFACTION_STABLE_AB1996:
            return self.best_solution_mip(
                enforce_stability=True, stability_ab1996=True
            )
        elif method == OptimizeMethod.MINIMIZE_DISSATISFACTION_STABLE_CUSTOM:
            return self.best_solution_mip(
                enforce_stability=True, stability_ab1996=False
            )
        elif method == OptimizeMethod.MINIMIZE_DISSATISFACTION_STABLE:
            return self.best_solution_mip(
                enforce_stability=True, stability_ab1996=True
            ) or self.best_solution_mip(
                enforce_stability=True, stability_ab1996=False
            )
        elif method == OptimizeMethod.MINIMIZE_DISSATISFACTION_STABLE_FALLBACK:
            solution = self.best_solution_mip(
                enforce_stability=True, stability_ab1996=True
            ) or self.best_solution_mip(
                enforce_stability=True, stability_ab1996=False
            )
            if solution:
                return solution
            log.warning(
                "Stable solution not found. Falling back to "
                "overall best (permitting instability)."
            )
            return self.best_solution_mip(enforce_stability=False)
        elif method == OptimizeMethod.ABRAHAM_STUDENT:
            return self.best_solution_abraham(optimal="student")
        elif method == OptimizeMethod.ABRAHAM_SUPERVISOR:
            return self.best_solution_abraham(optimal="supervisor")
        else:
            raise AssertionError(f"Unknown optimization method: {method!r}")

    # -------------------------------------------------------------------------
    # Solve via MIP
    # -------------------------------------------------------------------------

    def best_solution_mip(
        self, enforce_stability: bool = False, stability_ab1996: bool = False
    ) -> Optional[Solution]:
        """
        Return the best solution by optimizing with the MIP package.
        This is extremely impressive.
        See https://python-mip.readthedocs.io/.

        Args:
            enforce_stability:
                Ensure only stable "marriages" are produced (or fail entirely)?
            stability_ab1996:
                For ``enforce_stability``: use the stability constraint that is
                equation 4 of Abeledo & Blum (1996,
                https://doi.org/10.1016/0024-3795(95)00052-6).
        """
        # ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
        # Basic setup
        # ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
        supervisor_weight = self.config.supervisor_weight
        assert 0 <= supervisor_weight <= 1
        student_weight = 1 - supervisor_weight
        log.info(
            f"MIP approach: student_weight={student_weight}, "
            f"supervisor_weight={supervisor_weight}, "
            f"enforce_stability={enforce_stability}, "
            f"stability_ab1996={stability_ab1996}"
        )
        n_students = len(self.students)
        n_projects = len(self.projects)
        using_max_projects_per_supervisor = any(
            supervisor.max_n_projects is not None
            for supervisor in self.supervisors
        )

        # ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
        # Eligibility map
        # ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
        eligible = [
            [
                self.eligibility.is_eligible(student, project)
                for p, project in enumerate(self.projects)  # second index
            ]
            for s, student in enumerate(self.students)  # first index
        ]  # indexed s, p

        # ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
        # Dissatisfaction scores for each project/student combination
        # ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
        # CAUTION: get indexes the right way round!
        student_dissatisfaction_with_project = [
            [
                self.students[s].exponentiated_dissatisfaction(
                    self.projects[p]
                )
                for p in range(n_projects)  # second index
            ]
            for s in range(n_students)  # first index
        ]  # indexed s, p
        project_dissatisfaction_with_student = [
            [
                self.projects[p].exponentiated_dissatisfaction(
                    self.students[s]
                )
                for p in range(n_projects)  # second index
            ]
            for s in range(n_students)  # first index
        ]  # indexed s, p
        weighted_dissatisfaction = [
            [
                (
                    student_weight * student_dissatisfaction_with_project[s][p]
                    + supervisor_weight
                    * project_dissatisfaction_with_student[s][p]
                )
                for p in range(n_projects)  # second index
            ]
            for s in range(n_students)  # first index
        ]  # indexed s, p

        # ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
        # Model
        # ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
        m = Model("Student project allocation")
        # Binary variables to optimize, each linking a student to a project
        # CAUTION: get indexes the right way round!
        x = [
            [
                (
                    m.add_var(f"x[s={s},p={p}]", var_type=BINARY)
                    if eligible[s][p]
                    else None
                )
                for p in range(n_projects)  # second index
            ]
            for s in range(n_students)  # first index
        ]  # indexed s, p
        if using_max_projects_per_supervisor:
            # Create intermediate binary variables to indicate if a project is
            # in use (has been allocated), for certain projects (those whose
            # supervisors have capped the number of projects that they can
            # take). See below for explanation.
            project_in_use = [
                (
                    m.add_var(f"project_in_use[p={p}]", var_type=BINARY)
                    if self.projects[
                        p
                    ].at_least_one_supervisor_has_a_project_cap()
                    else None
                )  # don't bother for supervisors that don't care
                for p in range(n_projects)
            ]  # indexed: p
        else:
            project_in_use = []  # type: List[Var]

        # ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
        # Objective: happy students/supervisors
        # ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
        m.objective = minimize(
            xsum(
                x[s][p] * weighted_dissatisfaction[s][p]
                for p in range(n_projects)
                for s in range(n_students)
                if eligible[s][p]
            )
        )

        # ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
        # Constraint: For each student, exactly one project.
        # ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
        for s in range(n_students):
            m += (
                xsum(x[s][p] for p in range(n_projects) if eligible[s][p])
                == 1,
                f"student_{s}_one_project",
            )
            # Using a Special Ordered Set here doesn't materially speed things
            # up (maybe very slightly). I'm not entirely sure what the "weight"
            # parameter should be. Always 1? Or consecutive?
            # - https://docs.python-mip.com/en/latest/examples.html#exsos
            #   ... not terribly clear, but does use non-sequential order (of
            #   possible plants in a region) as weights.
            # - http://lpsolve.sourceforge.net/5.5/SOS.htm
            # - https://en.wikipedia.org/wiki/Special_ordered_set
            #   ... the benefit is for speed.
            # - https://www.tu-chemnitz.de/mathematik/discrete/manuals/cplex/doc/pdf/cplex81userman.pdf  # noqa
            #   ... gives an example (p244) using ordered warehouse size as the
            #   weights.
            # So:
            m.add_sos(
                sos=[
                    (x[s][p], p)  # p (the non-sequential order) is the weight
                    for p in range(n_projects)
                    if eligible[s][p]
                ],
                sos_type=1,  # Type 1: only one variable can receive value 1.
            )
        del s

        # ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
        # Constraint: For each project, up to the maximum number of students.
        # ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
        for p, project in enumerate(self.projects):
            m += (
                xsum(x[s][p] for s in range(n_students) if eligible[s][p])
                <= project.max_n_students,
                f"project_{p}_max_{project.max_n_students}_students",
            )
        del p, project

        # ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
        # Constraint: Maximum number of projects per supervisor
        # ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
        # This one is hard.
        # General advice on formulating these problems:
        # - http://people.brunel.ac.uk/~mastjjb/jeb/or/moreip.html
        # - https://pubsonline.informs.org/doi/pdf/10.1287/ited.2017.0177
        #   ^^^
        #   THIS ONE! Stevens & Palocsay (2017). Excellent.
        # Somewhat related problems:
        # - http://yetanothermathprogrammingconsultant.blogspot.com/2018/04/a-difficult-mip-construct-counting.html  # noqa
        # - https://math.stackexchange.com/questions/2732897/linear-integer-programming-count-consecutive-ones  # noqa
        if using_max_projects_per_supervisor:
            # - We work out whether each project is allocated, using SEPARATE
            #   (BINARY) VARIABLES.
            # - The rule for project p is:
            #       "If any student is allocated to p, then project_in_use[p]."
            # - That is:
            #       "SOME student allocated -> ALL that project in use."
            # - By the Decomposition rule, that translates to:
            #       student 1 allocated to p -> project p is in use
            #       student 2 allocated to p -> project p is in use
            #       ...
            # - By the Translation rule, each one can be represented by
            #       student_1_allocated_to_p <= project_p_in_use
            # - Converting that to a form with constants on the right,
            #       student_1_allocated_to_p - project_p_in_use <= 0

            for sv, supervisor in enumerate(self.supervisors):
                if supervisor.max_n_projects is not None:
                    # 1. Define whether relevant projects are in use.
                    for p in range(n_projects):
                        if not self.projects[p].is_supervised_by(supervisor):
                            continue
                        for s in range(n_students):
                            m += (
                                x[s][p] - project_in_use[p] <= 0,
                                f"project_{p}_in_use_by_student_{s}",
                            )
                    # 2. Constrain the number of projects for the supervisor.
                    m += (
                        xsum(
                            project_in_use[p]
                            for p in range(n_projects)
                            if self.projects[p].is_supervised_by(supervisor)
                        )
                        <= supervisor.max_n_projects,
                        f"supervisor_{sv}_max_{supervisor.max_n_projects}_projects",  # noqa
                    )
            del sv, supervisor

        # ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
        # Constraint: Maximum number of students per supervisor
        # ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
        for sv, supervisor in enumerate(self.supervisors):
            if supervisor.max_n_students is not None:
                m += (
                    xsum(
                        # "All students allocated to projects supervised by
                        # this supervisor."
                        x[s][p]
                        for s in range(n_students)
                        for p in range(n_projects)
                        if (
                            self.projects[p].is_supervised_by(supervisor)
                            and eligible[s][
                                p
                            ]  # don't consider impossible pairings
                        )
                    )
                    <= supervisor.max_n_students,
                    (
                        f"supervisor_{sv}_"
                        f"max_{supervisor.max_n_students}_students"
                    ),
                )
        del sv, supervisor

        # ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
        # Constraint: Only stable "marriages"?
        # ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
        if enforce_stability and stability_ab1996:
            # +++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
            # Stability via Abeledo & Blum 1996, assuming strict preferences
            # +++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
            log.info(
                "Trying for stability via Abeledo & Blum 1996, which "
                "assumes strict preferences."
            )
            if not self.are_preferences_strict_over_relevant_combos():
                log.error(
                    "Stability constraints of Abeledo & Blum (1996) require "
                    "strict preferences, but preferences are not strict (of "
                    "students for their projects, and of projects/supervisors "
                    "for all students who picked them). Failing."
                )
                return None
            # Equation 4 of Abeledo & Blum (1996), as above: the stability
            # constraint. We'll use their notation for clarity.
            # When they say a >{x} b, they mean "x prefers a to b".
            # Similarly, "a <{x} b" means "x prefers b to a".
            for u, v in product(range(n_students), range(n_projects)):
                # We'll say the student is "u" and the project is "v"
                if not eligible[u][v]:
                    continue
                student_dis = student_dissatisfaction_with_project[u][v]
                project_dis = project_dissatisfaction_with_student[u][v]
                other_project_vars = []  # type: List[Var]
                other_student_vars = []  # type: List[Var]
                for i in [_ for _ in range(n_projects) if _ != v]:  # "i"
                    if not eligible[u][i]:
                        continue
                    if (
                        student_dissatisfaction_with_project[u][i]
                        < student_dis
                    ):
                        # Student "u" prefers project "i" to project "v";
                        # that is, i >{u} v.
                        other_project_vars.append(x[u][i])
                for j in [_ for _ in range(n_students) if _ != u]:  # "j"
                    if not eligible[j][v]:
                        continue
                    if (
                        project_dissatisfaction_with_student[j][v]
                        < project_dis
                    ):
                        # Project "v" prefers student "j" to student "u";
                        # that is, j >{v} u.
                        other_student_vars.append(x[j][v])
                        # I'm pretty sure they must mean x{j,v} not x{v,j},
                        # since the variable x is always suffixed
                        # {u-type-thing, v-type-thing}, e.g. page 323.
                vars_to_sum = (
                    other_project_vars
                    + other_student_vars  # sum{for i >{u} v}{x{u,i}}
                    + [x[u][v]]  # sum{for j >{v} u}{x{j,v}}  # "x{u,v}"
                )
                stability_constraint = xsum(vars_to_sum) >= 1  # Eq. 4.
                log.debug(
                    f"Adding stability constraint: " f"{stability_constraint}"
                )
                m += stability_constraint, f"stability_s{u}_p{v}"
            # What's the logic here?
            # Lemma 3.1, which includes equation 4, is from ref. [2], which is
            # Abeledo & Rothblum (1994,
            # https://doi.org/10.1016/0166-218X(94)90130-9).
            # In that work, it's Theorem 3.1, Equation 7, p6. The logic is:
            # - If there is no better match for either the student or the
            #   project, then the first two components vanish, x{u,v} is the
            #   best match, and must be 1. [RNC CAVEAT: THAT REQUIRES STRICT
            #   ORDERING, which is one of the assumptions. So we need to deal
            #   with that.] So this seems to be saying "if there's no better
            #   match, pick it".
            # - Combined with the other inequalities, which say things like
            #   "everyone must be assigned" and "not too many students per
            #   project", forcing us to pick the best excludes picking anything
            #   that isn't the best.
            # - [In passing, note that we have a "bipartite" situation
            #   (projects are distinct from students; defined on p3).]
            del u, v
            del vars_to_sum, stability_constraint

        elif enforce_stability and not stability_ab1996:
            # +++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
            # Stability via a custom approach allowing non-strict preferences
            # +++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
            # Can we develop an equivalent when there might be indifference?
            # We want to say simply "if there's a better marriage, don't pick
            # this one".
            log.info(
                "Trying for stability via a custom method, which "
                "does not assume strict preferences. (Can be slow.)"
            )
            stability_constraints = set()  # type: Set[LinExpr]
            stability_constraint_tuples = []  # type: List[Tuple[LinExpr, str]]
            for s_idx, p_idx in product(range(n_students), range(n_projects)):
                if not eligible[s_idx][p_idx]:
                    continue
                s = self.students[s_idx]
                p = self.projects[p_idx]
                # So, for every eligible student/project combination...
                for other_p in self.gen_better_projects(s, p):
                    # other_p: "Other projects that s prefers to p."
                    other_p_idx = self.projects.index(other_p)
                    for other_s in self.gen_worse_students(other_p, s):
                        # other_s: "Other students that other_p would
                        # reject in favour of s."
                        other_s_idx = self.students.index(other_s)
                        constraint = (
                            # "Do not assign s to p and simultaneously
                            # assign other_s to other_p (because s and
                            # other_p would rather pair up with each
                            # other)." That is, s and other_p represent a
                            # blocking pair for a solution that includes a
                            # match between s and p and also between
                            # other_s and other_p.
                            x[s_idx][p_idx] + x[other_s_idx][other_p_idx]
                            <= 1
                            # You can't multiply these variables, but you
                            # can add them.
                        )
                        if constraint not in stability_constraints:
                            # We use a set because otherwise we may add the
                            # same thing several times.
                            log.debug(
                                f"Adding stability constraint: {constraint}, "
                                f"for s={s}, p={p}, "
                                f"other_s={other_s}, other_p={other_p}"
                            )
                            stability_constraints.add(constraint)
                            stability_constraint_tuples.append(
                                (
                                    constraint,
                                    f"stability_s{s_idx}_p{p_idx}_"
                                    f"other_s{other_s_idx}_"
                                    f"other_p{other_p_idx}",
                                )
                            )
                del s, p
            log.info(
                f"Adding {len(stability_constraints)} unique "
                f"stability constraints"
            )
            for stability_constraint_tuple in stability_constraint_tuples:
                m += stability_constraint_tuple
            del s_idx, p_idx
            del stability_constraints, stability_constraint_tuples

        # ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
        # Debug?
        # ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
        # import pdb; pdb.set_trace()
        if self.config.debug_model:
            report_on_model(m)

        # ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
        # Optimize
        # ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
        m.optimize(max_seconds=self.config.max_time_s)

        # ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
        # Debug?
        # ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
        if self.config.debug_model:
            report_on_model(m, solution_only=True)

        # ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
        # Extract results
        # ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
        if not m.num_solutions:
            return None
        # noinspection PyTypeChecker
        project_indexes = [
            next(
                p
                for p in range(n_projects)
                if eligible[s][p] and x[s][p].x >= ALMOST_ONE
            )
            # ... note that the value of a solved variable is var.x
            # If those two expressions are not the same, there's a bug.
            for s in range(n_students)
        ]
        solution = self._make_solution(project_indexes)
        if enforce_stability:
            assert solution.is_stable()
        return solution

    def _make_solution(
        self, project_indexes: Sequence[int], validate: bool = True
    ) -> Solution:
        """
        Creates a solution from project index numbers.

        Args:
            project_indexes:
                Indexes (zero-based) of project numbers, one per student,
                in the order of ``self.students``.
            validate:
                validate input? For debugging only.
        """
        if validate:
            n_students = len(self.students)
            assert (
                len(project_indexes) == n_students
            ), "Number of project indices does not match number of students"
        allocation = {}  # type: Dict[Student, Project]
        for student_idx, project_idx in enumerate(project_indexes):
            allocation[self.students[student_idx]] = self.projects[project_idx]
        return Solution(problem=self, allocation=allocation)

    # -------------------------------------------------------------------------
    # Solve via Abraham-Irving-Manlove 2007
    # -------------------------------------------------------------------------

    def best_solution_abraham(
        self, optimal: str = "student"
    ) -> Optional[Solution]:
        """
        Optimize via the Abraham-Irving-Manlove 2007 algorithm, optimally for
        students.

        For the Gale-Shapley algorithm, see

        - https://en.wikipedia.org/wiki/Gale%E2%80%93Shapley_algorithm
        - https://www.nrmp.org/nobel-prize/

        Others' work on Gale-Shapley:

        - https://towardsdatascience.com/gale-shapley-algorithm-simply-explained-caa344e643c2
        - https://gist.github.com/joyrexus/9967709 (a good one)
        - https://github.com/Vishal-Kancharla/Gale-Shapley-Algorithm
        - https://rosettacode.org/wiki/Stable_marriage_problem

        For Abraham-Irving-Manlove:

        - https://matching.readthedocs.io/
        """  # noqa
        assert optimal in ["student", "supervisor"]

        # ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
        # Set up the problem: (1) create objects
        # ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~

        mg_students = []  # type: List[MGStudent]
        student_to_mg_student = {}  # type: Dict[Student, MGStudent]
        mg_student_to_student = {}  # type: Dict[MGStudent, Student]
        for s in self.students:
            mg_student = MGStudent(name=s.name)
            mg_students.append(mg_student)
            student_to_mg_student[s] = mg_student
            mg_student_to_student[mg_student] = s

        mg_supervisors = []  # type: List[MGSupervisor]
        mg_projects = []  # type: List[MGProject]
        project_to_mg_supervisor = {}  # type: Dict[Project, MGSupervisor]
        mg_supervisor_to_project = {}  # type: Dict[MGSupervisor, Project]
        project_to_mg_project = {}  # type: Dict[Project, MGProject]
        mg_project_to_project = {}  # type: Dict[MGProject, Project]
        for p in self.projects:
            mg_supervisor = MGSupervisor(
                name=f"Supervisor of {p.title}", capacity=p.max_n_students
            )
            mg_supervisors.append(mg_supervisor)
            project_to_mg_supervisor[p] = mg_supervisor
            mg_supervisor_to_project[mg_supervisor] = p

            mg_project = MGProject(name=p.title, capacity=p.max_n_students)
            mg_project.set_supervisor(mg_supervisor)
            mg_projects.append(mg_project)
            project_to_mg_project[p] = mg_project
            mg_project_to_project[mg_project] = p

        # ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
        # Set up the problem: (2) define preferences
        # ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~

        # Student preferences. Must set these first.
        for s in self.students:
            mg_student = student_to_mg_student[s]
            preferred_projects = [
                project_to_mg_project[p]
                for p in s.projects_in_descending_order(
                    [
                        # Only the projects that the student has ranked
                        # explicitly...
                        px
                        for px in s.preferences.items_explicitly_ranked()
                        # ... and that the student is eligible for.
                        if self.eligibility.is_eligible(s, px)
                    ]
                )
            ]
            log.debug(
                f"For student {mg_student}, "
                f"setting preferences: {preferred_projects}"
            )
            mg_student.set_prefs(preferred_projects)

        # Supervisor/project preferences. (These are assigned to supervisors.)
        for p in self.projects:
            mg_supervisor = project_to_mg_supervisor[p]
            preferred_students = [
                student_to_mg_student[s]
                for s in p.students_in_descending_order(
                    [
                        # Only the students that explicitly chose this
                        # project...
                        sx
                        for sx in self.students_who_chose(p)
                        # ... and are eligible for it:
                        if self.eligibility.is_eligible(sx, p)
                    ]
                )
            ]
            log.debug(
                f"For supervisor {mg_supervisor}, "
                f"setting preferences: {preferred_students}"
            )
            mg_supervisor.set_prefs(preferred_students)

        # log.critical(f"Supervisors: {mg_supervisors}")
        # log.critical(f"Projects: {mg_projects}")
        # log.critical(f"Students: {mg_students}")

        # ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
        # Solve
        # ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
        game = MGStudentAllocation(mg_students, mg_projects, mg_supervisors)
        matching = game.solve(optimal=optimal)
        if matching is None:  # no solution found
            return None
        assert game.check_validity()
        assert game.check_stability()
        # log.critical(repr(matching))

        # ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
        # Translate back to our Solution class
        # ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
        allocation = {}  # type: Dict[Student, Project]
        for mg_project_copy, mg_student_copies in matching.items():
            # We cannot do a lookup. It has done a deepcopy. We have to match
            # by name.
            mg_project = next(
                mgp for mgp in mg_projects if mgp.name == mg_project_copy.name
            )
            project = mg_project_to_project[mg_project]
            for mg_student_copy in mg_student_copies:
                # Ditto... this is a bit silly...
                mg_student = next(
                    mgs
                    for mgs in mg_students
                    if mgs.name == mg_student_copy.name
                )
                student = mg_student_to_student[mg_student]
                allocation[student] = project
        # Create the solution
        solution = Solution(problem=self, allocation=allocation)
        # Sanity-check the solution
        unallocated_students = [
            s for s in self.students if s not in allocation
        ]
        if unallocated_students:
            log.critical(solution)
            log.critical(
                f"Failed: unallocated students: " f"{unallocated_students}"
            )
            return None
        return solution
