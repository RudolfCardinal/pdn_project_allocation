#!/usr/bin/env python

"""
pdn_project_allocation/pdn_project_allocation.py

See README.rst

Development notes
-----------------

Slightly tricky question: optimizing mean versus variance.

- Dissatisfaction mean: lower is better, all else being equal.
- Dissatisfaction variance: lower is better, all else being equal.
- So we have two options:

  - Optimize mean, then use variance as tie-breaker.
  - Optimize a weighted combination of mean and variance.

- Note that "least variance" itself is a rubbish proposition; that can mean
  "consistently bad".

- The choice depends whether greater equality can outweight slightly worse
  mean (dis)satisfaction. I've not found a good example of this. Optimizing
  mean happiness seems to be fine.

- Since moving to a MILP method (see below), we just optimize total weighted
  dissatisfaction.

Consistency:

- For the old brute-force approach, there was a need to break ties randomly,
  e.g. in the case of two projects with both students ranking the first project
  top. Consistency is important and lack of bias (e.g. alphabetical bias) is
  important, so we (a) set a consistent random number seed; (b)
  deterministically and then randomly sort the students; (c) run the optimizer.
  This gives consistent results and does not depend on e.g. alphabetical
  ordering, who comes first in the spreadsheet, etc. (No such effort is applied
  to project ordering.)

- The new MILP method also seems to be consistent.

AIM2007 troubles
----------------

.. code-block:: python

    from matching.games import StudentAllocation

    student_to_preferences = {
        "S1": ["P1", "P2", "P3"],
        "S2": ["P1", "P2", "P3"],
        "S3": ["P4", "P5", "P6"],
        "S4": ["P4", "P5", "P6"],
        "S5": ["P7", "P8", "P9"],
        "S6": ["P7", "P8", "P9"],
        "S7": ["P10", "P1", "P2"],
        "S8": ["P9", "P10", "P1"],
        "S9": ["P8", "P9", "P10"],
        "S10": ["P5", "P6", "P7"],
    }
    supervisor_to_preferences = {
        "SV1": ["S2", "S8", "S1", "S7"],
        "SV2": ["S2", "S1", "S7"],
        "SV3": ["S2", "S1"],
        "SV4": ["S3", "S4"],
        "SV5": ["S3", "S4", "S10"],
        "SV6": ["S3", "S4", "S10"],
        "SV7": ["S5", "S6", "S10"],
        "SV8": ["S5", "S6", "S9"],
        "SV9": ["S8", "S5", "S6", "S9"],
        "SV10": ["S8", "S9", "S7"],
    }
    project_to_supervisor = {
        "P1": "SV1",
        "P2": "SV2",
        "P3": "SV3",
        "P4": "SV4",
        "P5": "SV5",
        "P6": "SV6",
        "P7": "SV7",
        "P8": "SV8",
        "P9": "SV9",
        "P10": "SV10",
    }
    project_to_capacity = {
        "P1": 1,
        "P2": 1,
        "P3": 1,
        "P4": 1,
        "P5": 1,
        "P6": 1,
        "P7": 1,
        "P8": 1,
        "P9": 1,
        "P10": 1,
    }
    supervisor_to_capacity = {
        "SV1": 1,
        "SV2": 1,
        "SV3": 1,
        "SV4": 1,
        "SV5": 1,
        "SV6": 1,
        "SV7": 1,
        "SV8": 1,
        "SV9": 1,
        "SV10": 1,
    }

    game = StudentAllocation.create_from_dictionaries(
        student_to_preferences,
        supervisor_to_preferences,
        project_to_supervisor,
        project_to_capacity,
        supervisor_to_capacity,
    )

    matching = game.solve(optimal="student")
    assert game.check_validity()  # OK
    assert game.check_stability()  # OK

    # But, what it doesn't tell you:

    print(matching)

    # {P1: [S2], P2: [S1], P3: [], P4: [S3], P5: [S4], P6: [S10], P7: [S5], P8: [S6], P9: [S8], P10: [S9]}
    # ... i.e. P3 has no student, and S7 has no project.

"""  # noqa

import argparse
from collections import OrderedDict
import csv
import datetime
from enum import Enum
from itertools import product
import logging
import operator
import os
import random
from statistics import mean, median, variance
import sys
import traceback
from typing import (Any, Dict, Generator, List, Optional, Sequence, Set,
                    Tuple, Union)

from cardinal_pythonlib.argparse_func import RawDescriptionArgumentDefaultsHelpFormatter  # noqa
from cardinal_pythonlib.enumlike import (
    CaseInsensitiveEnumMeta,
    keys_descriptions_from_enum,
)
from cardinal_pythonlib.logs import main_only_quicksetup_rootlogger
from cardinal_pythonlib.maths_py import sum_of_integers_in_inclusive_range
from cardinal_pythonlib.cmdline import cmdline_quote
from cardinal_pythonlib.reprfunc import auto_repr
from openpyxl.cell import Cell
from openpyxl.reader.excel import load_workbook
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from matching.games.student_allocation import (
    Project as MGProject,
    Student as MGStudent,
    StudentAllocation as MGStudentAllocation,
    Supervisor as MGSupervisor,
)
from mip import BINARY, LinExpr, minimize, Model, Var, xsum
from mip.exceptions import SolutionNotAvailable

log = logging.getLogger(__name__)

VERSION = "1.2.0"
VERSION_DATE = "2020-10-05"

ALMOST_ONE = 0.99
DEFAULT_PREFERENCE_POWER = 1.0
DEFAULT_MAX_SECONDS = 1e100  # the default in mip
DEFAULT_SUPERVISOR_WEIGHT = 0.3  # 70% student, 30% supervisor by default
RNG_SEED = 1234  # fixed
VERY_VERBOSE = False  # debugging option

EXT_XLSX = ".xlsx"
EXIT_FAILURE = 1
EXIT_SUCCESS = 0

INPUT_TYPES_SUPPORTED = [EXT_XLSX]
OUTPUT_TYPES_SUPPORTED = INPUT_TYPES_SUPPORTED

TRUE_VALUES = [1, "Y", "y", "T", "t"]
FALSE_VALUES = [0, "N", "n", "F", "f"]
MISSING_VALUES = ["", None]


# =============================================================================
# Playing with the mip package
# =============================================================================

r"""

Just for fun, the n-queens problem from
https://python-mip.readthedocs.io/en/latest/examples.html:

.. code-block:: python

from sys import stdout
from mip import Model, xsum, MAXIMIZE, BINARY

# number of queens
n = 60

queens = Model()

x = [[queens.add_var(f"x({i},{j})", var_type=BINARY)
      for j in range(n)] for i in range(n)]

# one per row
for i in range(n):
    queens += xsum(x[i][j] for j in range(n)) == 1, f"row({i})"

# one per column
for j in range(n):
    queens += xsum(x[i][j] for i in range(n)) == 1, f"col({j})"

# diagonal \
for p, k in enumerate(range(2 - n, n - 2 + 1)):
    queens += xsum(x[i][j] for i in range(n) for j in range(n)
                   if i - j == k) <= 1, f"diag1({p})"

# diagonal /
for p, k in enumerate(range(3, n + n)):
    queens += xsum(x[i][j] for i in range(n) for j in range(n)
                   if i + j == k) <= 1, f"diag2({p})"

queens.optimize()

text = ""
if queens.num_solutions:
    for i, v in enumerate(queens.vars):
        text += 'Q ' if v.x >= 0.99 else '. '
        if i % n == n-1:
            text += "\n"

print(text)
# for v in queens.vars: print(v)
# for c in queens.constrs: print(c)
# print(queens.objective)  # blank

"""


# =============================================================================
# Enum classes
# =============================================================================

class SheetNames(object):
    """
    Sheet names within the input/output spreadsheet file.
    """
    ELIGIBILITY = "Eligibility"
    INFORMATION = "Information"  # output
    PROJECT_POPULARITY = "Project_popularity"  # output
    PROJECT_ALLOCATIONS = "Project_allocations"  # output
    PROJECTS = "Projects"  # input, output
    STUDENT_ALLOCATIONS = "Student_allocations"  # output
    STUDENT_PREFERENCES = "Student_preferences"  # input, output
    SUPERVISORS = "Supervisors"  # input, output
    SUPERVISOR_PREFERENCES = "Supervisor_preferences"  # input, output


class SheetHeadings(object):
    """
    Column headings within the input spreadsheet.
    """
    MAX_NUMBER_OF_PROJECTS = "Max_number_of_projects"
    MAX_NUMBER_OF_STUDENTS = "Max_number_of_students"
    PROJECT = "Project"
    SUPERVISOR = "Supervisor"


class OptimizeMethod(Enum, metaclass=CaseInsensitiveEnumMeta):
    MINIMIZE_DISSATISFACTION = "Minimize weighted dissatisfaction"
    MINIMIZE_DISSATISFACTION_STABLE_AB1996 = (
        "Minimize weighted dissatisfaction, requiring stability, "
        "via Abeledo & Blum (1996) method"
    )
    MINIMIZE_DISSATISFACTION_STABLE_CUSTOM = (
        "Minimize weighted dissatisfaction, requiring stability, "
        "via custom method that does not assume strict preferences"
    )
    MINIMIZE_DISSATISFACTION_STABLE = (
        "Minimize weighted dissatisfaction, requiring stability, "
        "via Abeledo & Blum (1996) falling back to custom method if required"
    )
    MINIMIZE_DISSATISFACTION_STABLE_FALLBACK = (
        "Minimize weighted dissatisfaction, requiring stability if possible"
        "(as for MINIMIZE_DISSATISFACTION_STABLE), but falling back to "
        "unstable if not."
    )
    ABRAHAM_STUDENT = "Abraham-Irving-Manlove 2007 (optimal for students)"
    ABRAHAM_SUPERVISOR = (
        "Abraham-Irving-Manlove 2007 (optimal for supervisors)"
    )


DEFAULT_METHOD = OptimizeMethod.MINIMIZE_DISSATISFACTION_STABLE_FALLBACK


# =============================================================================
# Helper functions
# =============================================================================

def mismatch(actual: List[Any], expected: List[Any]) -> str:
    """
    Provides text to locate a mismatch between two lists.
    """
    n_actual = len(actual)
    n_intended = len(expected)
    if n_actual != n_intended:
        return (
            f"Wrong length: actual has length {n_actual}, "
            f"intended has length {n_intended}"
        )
    for i in range(n_actual):
        if actual[i] != expected[i]:
            return f"Found {actual[i]!r} where {expected[i]!r} was expected"
    return ""


def is_empty_row(row: Sequence[Cell]) -> bool:
    """
    Is this an empty spreadsheet row?
    """
    return all(cell.value is None for cell in row)


def read_until_empty_row(ws: Worksheet) -> List[List[Any]]:
    """
    Reads a spreadsheet until the first empty line.
    (Helpful because Excel spreadsheets are sometimes seen as having 1048576
    rows when they don't really).
    """
    rows = []  # type: List[List[Any]]
    for row in ws.iter_rows():
        if is_empty_row(row):
            break
        rows.append([cell.value for cell in row])
    return rows


def report_on_model(m: Model,
                    loglevel: int = logging.WARNING,
                    solution_only: bool = False) -> None:
    """
    Shows detail of a MIP model to the log.
    """
    lines = ["Model:", "", "- Variables:", ""]
    try:
        for v in m.vars:
            lines.append(f"{v.name} == {v.x}")
    except SolutionNotAvailable:
        if solution_only:
            raise
        for v in m.vars:
            lines.append(f"{v.name}")
    if not solution_only:
        lines += ["", "- Objective:", ""]
        lines.append(str(m.objective.sense))
        lines.append(str(m.objective))
        lines += ["", "- Constraints:", ""]
        for c in m.constrs:
            lines.append(str(c))
    log.log(loglevel, "\n".join(lines))


# =============================================================================
# Master config
# =============================================================================

class Config(object):
    """
    Master config object.
    """

    def __init__(
            self,
            filename: str,
            allow_defunct_projects: bool = False,
            allow_student_preference_ties: bool = False,
            allow_supervisor_preference_ties: bool = False,
            cmd_args: Dict[str, Any] = None,
            debug_model: bool = False,
            max_time_s: float = DEFAULT_MAX_SECONDS,
            missing_eligibility: bool = None,
            no_shuffle: bool = False,
            optimize_method: OptimizeMethod = DEFAULT_METHOD,
            preference_power: float = DEFAULT_PREFERENCE_POWER,
            student_must_have_choice: bool = False,
            supervisor_weight: float = DEFAULT_SUPERVISOR_WEIGHT) -> None:
        """
        Reads a file, autodetecting its format, and returning the
        :class:`Problem`.

        Args:
            filename:
                Source data file to read.

            allow_defunct_projects:
                Allow projects that permit no students?
            allow_student_preference_ties:
                Allow students to express preference ties?
            allow_supervisor_preference_ties:
                Allow supervisors to express preference ties?
            cmd_args:
                Copy of command-line arguments
            debug_model:
                Report the MIP model before solving it?
            max_time_s:
                Time limit for MIP optimizer (s).
            missing_eligibility:
                Use ``True`` or ``False`` to treat missing eligibility cells
                as meaning "eligible" or "ineligible", respectively, or
                ``None`` to treat blank cells as invalid.
            no_shuffle:
                Don't shuffle anything. FOR DEBUGGING ONLY.
            optimize_method:
                Method to use for optimizing.
            preference_power:
                Power (exponent) to raise preferences to.
            student_must_have_choice:
                Prevent students being allocated to projects they've not
                explicitly ranked?
            supervisor_weight:
                Weight allocated to supervisor preferences; range [0, 1].
                (Student preferences are weighted as 1 minus this.)
        """
        self.filename = filename

        self.allow_defunct_projects = allow_defunct_projects
        self.allow_student_preference_ties = allow_student_preference_ties
        self.allow_supervisor_preference_ties = allow_supervisor_preference_ties  # noqa
        self.debug_model = debug_model
        self.max_time_s = max_time_s
        self.missing_eligibility = missing_eligibility
        self.no_shuffle = no_shuffle
        self.optimize_method = optimize_method
        self.preference_power = preference_power
        self.student_must_have_choice = student_must_have_choice
        self.supervisor_weight = supervisor_weight

        self.cmd_args = cmd_args

    def __str__(self) -> str:
        return str(self.cmd_args)


# =============================================================================
# Preferences
# =============================================================================

class Preferences(object):
    """
    Represents preference as a mapping from arbitrary objects (being preferred)
    to ranks.
    """
    def __init__(self,
                 n_options: int,
                 preferences: Dict[Any, Union[int, float]] = None,
                 owner: Any = None,
                 allow_ties: bool = False,
                 preference_power: float = DEFAULT_PREFERENCE_POWER) -> None:
        """
        Args:
            n_options:
                Total number of things to be judged.
            preferences:
                Mapping from "thing being judged" to "rank preference" (1 is
                best). If ``allow_ties`` is set, allows "2.5" for "joint
                second/third"; otherwise, they must be integer.
            owner:
                Person/thing expressing preferences (for cosmetic purposes
                only).
            allow_ties:
                Allows ties to be expressed.
            preference_power:
                Power (exponent) to raise preferences to.

        Other attributes:
        - ``available_dissatisfaction``: sum of [1 ... ``n_options`]
        - ``allocated_dissatisfaction``: sum of expressed preference ranks.
          (For example, if you only pick your top option, with rank 1, then you
          have expressed a total dissatisfaction of 1. If you have expressed
          a preference for rank #1 and rank #2, you have expressed a total
          dissatisfaction of 3.)
        """
        self._n_options = n_options
        self._preferences = OrderedDict()  # type: Dict[Any, Union[int, float]]
        self._owner = owner
        self._total_dissatisfaction = sum_of_integers_in_inclusive_range(
            1, n_options)
        self._allocated_dissatisfaction = 0
        self._allow_ties = allow_ties
        self._preference_power = preference_power

        if preferences:
            for item, rank in preferences.items():
                if rank is not None:
                    self.add(item, rank, _validate=False)
                    # ... defer validation until all data in...
            self._validate()  # OK, now validate

    def __str__(self) -> str:
        """
        String representation.
        """
        parts = ", ".join(f"{k} → {v}" for k, v in self._preferences.items())
        return (
            f"Preferences({parts}; "
            f"unranked options score {self._unranked_item_dissatisfaction})"
        )

    def __repr__(self) -> str:
        return "{" + ", ".join(
            f"{str(k)}: {str(v)}" for k, v in self._preferences.items()
        ) + "}"

    def set_n_options(self, n_options: int) -> None:
        """
        Sets the total number of options, and ensures that the preferences
        are compatible with this.
        """
        self._n_options = n_options
        self._validate()

    def add(self, item: Any, rank: float, _validate: bool = True) -> None:
        """
        Add a preference for an item.

        Args:
            item:
                Thing for which a preference is being assessed.
            rank:
                Integer preference rank (1 best, 2 next, etc.).
            _validate:
                Validate immediately?
        """
        if not self._allow_ties:
            assert item not in self._preferences, (
                f"Can't add same item twice (when allow_ties is False); "
                f"attempt to re-add {item!r}"
            )
            assert isinstance(rank, int), (
                f"Only integer preferences allowed "
                f"(when allow_ties is False); was {rank!r}"
            )
            assert rank not in self._preferences.values(), (
                f"No duplicate dissatisfaction scores allowed (when "
                f"allow_ties is False)): attempt to re-add rank {rank}"
            )
        self._preferences[item] = rank
        self._allocated_dissatisfaction += rank
        if _validate:
            self._validate()

    def _validate(self) -> None:
        """
        Validates:

        - that there are some options;
        - that preferences for all options are in the range [1, ``n_options``];
        - that the ``allocated_dissatisfaction`` is no more than the
          ``available_dissatisfaction``.

        Raises:
            :exc:`AssertionError` upon failure.
        """
        assert self._n_options > 0, "No options"
        for rank in self._preferences.values():
            assert 1 <= rank <= self._n_options, (
                f"Invalid preference: {rank!r} "
                f"(must be in range [1, {self._n_options}]"
            )
        n_expressed = len(self._preferences)
        expected_allocation = sum_of_integers_in_inclusive_range(
            1, n_expressed)
        assert self._allocated_dissatisfaction == expected_allocation, (
            f"Dissatisfaction scores add up to "
            f"{self._allocated_dissatisfaction}, but must add up to "
            f"{expected_allocation}, since you have expressed "
            f"{n_expressed} preferences (you can only express the 'top n' "
            f"preferences)"
        )
        assert (
            self._allocated_dissatisfaction <= self._total_dissatisfaction
        ), (
            f"Dissatisfaction scores add up to "
            f"{self._allocated_dissatisfaction}, which is more than the "
            f"maximum available of {self._total_dissatisfaction} "
            f"(for {self._n_options} options)"
        )

    @property
    def _unallocated_dissatisfaction(self) -> int:
        """
        The amount of available "dissatisfaction", not yet allocated to an
        item (see :class:`Preferences`).
        """
        return self._total_dissatisfaction - self._allocated_dissatisfaction

    @property
    def _unranked_item_dissatisfaction(self) -> Optional[float]:
        """
        The mean "dissatisfaction" (see :class:`Preferences`) for every option
        without an explicit preference, or ``None`` if there are no such
        options.
        """
        n_unranked = self._n_options - len(self._preferences)
        return (
            self._unallocated_dissatisfaction / n_unranked
            if n_unranked > 0 else None
        )

    def preference(self, item: Any) -> Union[int, float]:
        """
        Returns a numerical preference score for an item. Will use the
        "unranked" item dissatisfaction if no preference has been expressed for
        this particular item.

        Raises the raw preference score to ``preference_power`` (by default 1).

        Args:
            item:
                The item to look up.
        """
        return self._preferences.get(item, self._unranked_item_dissatisfaction)

    def exponentiated_preference(self, item: Any) -> Union[int, float]:
        """
        As for :meth:`preference`, but raised to ``preference_power`` (by
        default 1).

        Args:
            item:
                The item to look up.
        """
        return self.preference(item) ** self._preference_power

    def raw_preference(self, item: Any) -> Optional[int]:
        """
        Returns the raw preference for an item (for reproducing the input).

        Args:
            item:
                The item to look up.
        """
        return self._preferences.get(item)  # returns None if absent

    def actively_expressed_preference_for(self, item: Any) -> bool:
        """
        Did the person actively express a preference for this item?
        """
        return item in self._preferences

    def items_explicitly_ranked(self) -> List[Any]:
        """
        All the items for which there is an explicit preference.
        """
        return list(self._preferences.keys())

    def items_descending_order(
            self, all_items: List[Any]) -> List[Any]:
        """
        Returns all the items provided, in descending preference order (or the
        order provided, as a tie-break).
        """
        options = []  # type: List[Tuple[Any, float, int]]
        for i, item in enumerate(all_items):
            preference = self.preference(item)
            options.append((item, preference, i))
        return [
            t[0]  # the item
            for t in sorted(options, key=operator.itemgetter(1, 2))
        ]
        # ... sort by ascending dissatisfaction score (= descending
        # preference), then ascending sequence order

    def is_strict_over(self, items: List[Any]) -> bool:
        """
        Are all preferences strictly ordered for the items in question?
        """
        prefs = [self.preference(item) for item in items]
        n_preferences = len(prefs)
        n_unique_prefs = len(set(prefs))
        return n_preferences == n_unique_prefs

    def is_strict_over_expressed_preferences(self) -> bool:
        """
        Are preferences strictly ordered for the items for which a preference
        has been expressed?
        """
        return self.is_strict_over(self.items_explicitly_ranked())


# =============================================================================
# Student
# =============================================================================

class Student(object):
    """
    Represents a single student, with their preferences.
    """
    def __init__(self,
                 name: str,
                 number: int,
                 preferences: Dict["Project", int],
                 n_projects: int,
                 allow_ties: bool = False,
                 preference_power: float = DEFAULT_PREFERENCE_POWER) -> None:
        """
        Args:
            name:
                Student's name.
            number:
                Row number of student (cosmetic only).
            preferences:
                Map from project to rank preference (1 to ``n_projects``
                inclusive).
            n_projects:
                Total number of projects (for validating inputs).
            allow_ties:
                Allow ties in preferences?
            preference_power:
                Power (exponent) to raise preferences to.
        """
        self.name = name
        self.number = number
        self.preferences = Preferences(
            n_options=n_projects,
            preferences=preferences,
            owner=self,
            allow_ties=allow_ties,
            preference_power=preference_power
        )

    def __str__(self) -> str:
        """
        String representation.
        """
        return f"{self.name} (St#{self.number})"

    def __repr__(self) -> str:
        return auto_repr(self)

    def description(self) -> str:
        """
        Verbose description.
        """
        return f"{self}: {self.preferences}"

    def shortname(self) -> str:
        """
        Name and number.
        """
        return f"{self.name} (St#{self.number})"

    def __lt__(self, other: "Student") -> bool:
        """
        Comparison for sorting, used for console display.
        Default sort is by case-insensitive name.
        """
        return self.name.lower() < other.name.lower()

    def dissatisfaction(self, project: "Project") -> float:
        """
        How dissatisfied is this student if allocated a particular project?
        """
        return self.preferences.preference(project)

    def exponentiated_dissatisfaction(self, project: "Project") -> float:
        """
        As for :meth:`dissatisfaction`, but raised to the desired power.
        """
        return self.preferences.exponentiated_preference(project)

    def explicitly_ranked_project(self, project: "Project") -> bool:
        """
        Did the student explicitly rank this project?
        """
        return self.preferences.actively_expressed_preference_for(project)

    def projects_in_descending_order(
            self, all_projects: List["Project"]) -> List["Project"]:
        """
        Returns projects in descending order of preference.
        """
        return self.preferences.items_descending_order(all_projects)


# =============================================================================
# Supervisor
# =============================================================================

class Supervisor(object):
    """
    Simple representation of a supervisor.
    """
    def __init__(self,
                 name: str,
                 number: int,
                 max_n_projects: int = None,
                 max_n_students: int = None) -> None:
        """
        Args:
            name:
                Supervisor name.
            number:
                Supervisor number (cosmetic only: matches input order).
            max_n_projects:
                Maximum number of projects this supervisor can supervise.
                (They may offer more projects, but be unable to support all of
                them simultaneously.)
            max_n_students:
                Maximum number of students this supervisor can take.
        """
        assert name, "Missing supervisor name"
        assert number >= 1, "Bad supervisor number"
        assert max_n_projects is None or max_n_projects >= 1, (
            f"Supervisor {name!r}: invalid max_n_projects; must be None or "
            f">=1 but is {max_n_projects!r}"
        )
        assert max_n_students is None or max_n_students >= 1, (
            f"Supervisor {name!r}: invalid max_n_students; must be None or "
            f">=1 but is {max_n_projects!r}"
        )
        self.name = name
        self.number = number
        self.max_n_projects = max_n_projects
        self.max_n_students = max_n_students

    def __str__(self) -> str:
        """
        String representation.
        """
        return f"{self.name} (Sv#{self.number})"

    def __repr__(self) -> str:
        return auto_repr(self)

    def description(self) -> str:
        """
        Verbose description.
        """
        return (
            f"{self}: max_n_projects={self.max_n_projects}, "
            f"max_n_students={self.max_n_students}"
        )


# =============================================================================
# Project
# =============================================================================

class Project(object):
    """
    Simple representation of a project.
    """
    def __init__(self,
                 name: str,
                 number: int,
                 supervisor: Supervisor,
                 max_n_students: int,
                 allow_defunct_projects: bool = False) -> None:
        """
        Args:
            name:
                Project name.
            number:
                Project number (cosmetic only; matches input order).
            supervisor:
                The project's supervisor
            max_n_students:
                Maximum number of students supported.
            allow_defunct_projects:
                Allow projects that permit no students?
        """
        assert name, "Missing project name"
        assert number >= 1, "Bad project number"
        if allow_defunct_projects:
            assert max_n_students >= 0, "Bad max_n_students"
        else:
            assert max_n_students >= 1, "Bad max_n_students"
        self.name = name
        self.number = number
        self.supervisor = supervisor
        self.max_n_students = max_n_students
        self.supervisor_preferences = None  # type: Optional[Preferences]
        # ... the project supervisor's preferences for students with respect
        #     to THIS project.

    def __str__(self) -> str:
        """
        String representation.
        """
        return f"{self.name} (P#{self.number})"

    def __repr__(self) -> str:
        return auto_repr(self)

    def __lt__(self, other: "Project") -> bool:
        """
        Comparison for sorting, used for console display.
        Default sort is by case-insensitive name.
        """
        return self.name.lower() < other.name.lower()

    def description(self) -> str:
        """
        Describes the project.
        """
        return (
            f"{self} (max {self.max_n_students} students): "
            f"{self.supervisor_preferences}"
        )

    def set_supervisor_preferences(
            self,
            n_students: int,
            preferences: Dict[Student, int],
            allow_ties: bool = False,
            preference_power: float = DEFAULT_PREFERENCE_POWER) -> None:
        """
        Sets the supervisor's preferences about students for a project.
        """
        self.supervisor_preferences = Preferences(
            n_options=n_students,
            owner=self,
            preferences=preferences,
            allow_ties=allow_ties,
            preference_power=preference_power
        )

    def dissatisfaction(self, student: Student) -> float:
        """
        How dissatisfied is this project's supervisor if allocated a particular
        student?
        """
        return self.supervisor_preferences.preference(student)

    def exponentiated_dissatisfaction(self, project: "Project") -> float:
        """
        As for :meth:`dissatisfaction`, but raised to the desired power.
        """
        return self.supervisor_preferences.exponentiated_preference(project)

    def students_in_descending_order(
            self, all_students: List[Student]) -> List[Student]:
        """
        Returns students in descending order of preference.
        """
        return self.supervisor_preferences.items_descending_order(all_students)

    def is_supervised_by(self, supervisor: Supervisor) -> bool:
        """
        Is this the supervisor of this project?
        """
        return self.supervisor == supervisor


# =============================================================================
# Solution
# =============================================================================

class Solution(object):
    """
    Represents a potential solution.
    """
    def __init__(self,
                 problem: "Problem",
                 allocation: Dict[Student, Project]) -> None:
        """
        Args:
            problem:
                The :class:`Problem`, defining projects and students.
            allocation:
                The mapping of students to projects.
        """
        self.problem = problem
        self.allocation = allocation

    # -------------------------------------------------------------------------
    # Representations
    # -------------------------------------------------------------------------

    def __str__(self) -> str:
        """
        String representation.
        """
        lines = ["Solution:"]
        for student, project in self._gen_student_project_pairs():
            std = student.dissatisfaction(project)
            svd = project.dissatisfaction(student)
            lines.append(
                f"{student.shortname()} -> {project} "
                f"(student dissatisfaction {std}; "
                f"supervisor dissatisfaction {svd})")
        return "\n".join(lines)

    def shortdesc(self) -> str:
        """
        Very short description. Ordered by student number.
        """
        students = sorted(self.allocation.keys(), key=lambda s: s.number)
        parts = [f"{s.number}: {self.allocation[s].number}"
                 for s in students]
        return (
            "{" + ", ".join(parts) + "}" +
            f", student dissatisfaction {self.student_dissatisfaction_scores()}"
        )

    # -------------------------------------------------------------------------
    # Allocations
    # -------------------------------------------------------------------------

    def allocated_project(self, student: Student) -> Project:
        """
        Which project was allocated to this student?
        """
        return self.allocation[student]

    def allocated_students(self, project: Project) -> List[Student]:
        """
        Which students were allocated to this project?
        """
        return sorted(k for k, v in self.allocation.items() if v == project)

    def is_allocated(self, student: Student, project: Project) -> bool:
        """
        Is this student allocated to this project?
        """
        return self.allocation[student] == project

    def _gen_student_project_pairs(self) -> Generator[Tuple[Student, Project],
                                                      None, None]:
        """
        Generates ``student, project`` pairs in student order.
        """
        students = sorted(self.allocation.keys(), key=lambda s: s.number)
        for student in students:
            project = self.allocation[student]
            yield student, project

    # -------------------------------------------------------------------------
    # Student dissatisfaction
    # -------------------------------------------------------------------------

    def student_dissatisfaction_scores(self) -> List[float]:
        """
        All dissatisfaction scores.
        """
        dscores = []  # type: List[float]
        for student in self.problem.students:
            project = self.allocation[student]
            dscores.append(student.dissatisfaction(project))
        return dscores

    def student_dissatisfaction_median(self) -> float:
        """
        Median dissatisfaction per student.
        """
        return median(self.student_dissatisfaction_scores())

    def student_dissatisfaction_mean(self) -> float:
        """
        Mean dissatisfaction per student.
        """
        return mean(self.student_dissatisfaction_scores())

    def student_dissatisfaction_variance(self) -> float:
        """
        Variance of dissatisfaction scores.
        """
        return variance(self.student_dissatisfaction_scores())

    def student_dissatisfaction_min(self) -> float:
        """
        Minimum of dissatisfaction scores.
        """
        return min(self.student_dissatisfaction_scores())

    def student_dissatisfaction_max(self) -> float:
        """
        Maximum of dissatisfaction scores.
        """
        return max(self.student_dissatisfaction_scores())

    # -------------------------------------------------------------------------
    # Supervisor dissatisfaction
    # -------------------------------------------------------------------------

    def supervisor_dissatisfaction_scores_sum_students(self) -> List[float]:
        """
        All dissatisfaction scores. (If a project has several students, it
        scores the SUM of its dissatisfaction for each of those students
        scores.)
        """
        dscores = []  # type: List[float]
        for project in self.problem.projects:
            dscore = 0
            for student in self.problem.students:
                if self.allocation[student] == project:
                    dscore += project.dissatisfaction(student)
            dscores.append(dscore)
        return dscores

    def supervisor_dissatisfaction_scores_each_student(self) -> List[float]:
        """
        All dissatisfaction scores. (If a project has several students,
        multiple scores are returned for that project.)
        """
        dscores = []  # type: List[float]
        for project in self.problem.projects:
            for student in self.problem.students:
                if self.allocation[student] == project:
                    dscores.append(project.dissatisfaction(student))
        return dscores

    def supervisor_dissatisfaction_median(self) -> float:
        """
        Median dissatisfaction per student.
        """
        return median(self.supervisor_dissatisfaction_scores_each_student())

    def supervisor_dissatisfaction_mean(self) -> float:
        """
        Mean dissatisfaction per student.
        """
        return mean(self.supervisor_dissatisfaction_scores_each_student())

    def supervisor_dissatisfaction_variance(self) -> float:
        """
        Variance of dissatisfaction scores.
        """
        return variance(self.supervisor_dissatisfaction_scores_each_student())

    def supervisor_dissatisfaction_min(self) -> float:
        """
        Minimum of dissatisfaction scores.
        """
        return min(self.supervisor_dissatisfaction_scores_each_student())

    def supervisor_dissatisfaction_max(self) -> float:
        """
        Maximum of dissatisfaction scores.
        """
        return max(self.supervisor_dissatisfaction_scores_each_student())

    # -------------------------------------------------------------------------
    # Stability test
    # -------------------------------------------------------------------------

    def gen_better_projects(
            self,
            student: Student,
            project: Project) -> Generator[Project, None, None]:
        """
        Generates projects that this student prefers over the specified one.
        """
        for p in self.problem.gen_better_projects(student, project):
            yield p

    def gen_better_students(
            self,
            project: Project,
            student: Student) -> Generator[Student, None, None]:
        """
        Generates students that this project prefers over the specified one,
        for which they're eligible, AND who are are not already allocated to
        that project (bearing in mind that a project can have several
        students).
        """
        for s in self.problem.gen_better_students(project, student):
            if not self.is_allocated(s, project):
                yield s

    def stability(self, describe_all_failures: bool = True) -> Tuple[bool, str]:
        """
        Is the solution a stable match, and if not, why not? See README.rst for
        discussion. See also https://gist.github.com/joyrexus/9967709.

        Arguments:
            describe_all_failures:
                Show all reasons for failure.

        Returns:
            tuple: (stable, reason_for_instability)
        """
        stable = True
        instability_reasons = []  # type: List[str]
        for student, project in self._gen_student_project_pairs():
            for alt_project in self.gen_better_projects(student, project):
                for alt_proj_student in self.allocated_students(alt_project):
                    if alt_proj_student == student:
                        continue
                    if student in self.gen_better_students(alt_project,
                                                           alt_proj_student):
                        instability_reasons.append(
                            f"Pairing of student {student} to project "
                            f"{project} is unstable. "
                            f"The student would rather have alternative "
                            f"project {alt_project}, and that alternative "
                            f"project would rather have {student} than their "
                            f"current allocation of {alt_proj_student}."
                        )
                        stable = False
                        if not describe_all_failures:
                            return False, "\n\n".join(instability_reasons)
        if stable:
            return True, "[Stable]"
        else:
            return False, "\n\n".join(instability_reasons)

    def is_stable(self) -> bool:
        """
        Is the solution a stable match?
        """
        return self.stability(describe_all_failures=False)[0]

    # -------------------------------------------------------------------------
    # Saving
    # -------------------------------------------------------------------------

    def write_xlsx(self, filename: str) -> None:
        """
        Writes the solution to an Excel XLSX file (and its problem, for data
        safety).

        Args:
            filename:
                Name of file to write.
        """
        log.info(f"Writing output to: {filename}")
        wb = Workbook(write_only=True)  # doesn't create default sheet

        # ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
        # Allocations, by student
        # ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
        ss = wb.create_sheet(SheetNames.STUDENT_ALLOCATIONS)
        ss.append([
            "Student",
            "Project",
            "Supervisor",
            "Student's rank of (dissatisfaction with) allocated project",
        ])
        for student, project in self._gen_student_project_pairs():
            ss.append([
                student.name,
                project.name,
                project.supervisor.name,
                student.dissatisfaction(project),
            ])

        # ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
        # Allocations, by project
        # ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
        ps = wb.create_sheet(SheetNames.PROJECT_ALLOCATIONS)
        ps.append([
            "Project",
            "Supervisor",
            "Student(s)",
            "Students' rank(s) of (dissatisfaction with) allocated project",
            "Project supervisor's rank(s) of (dissatisfaction with) allocated student(s)",  # noqa
        ])
        for project in self.problem.sorted_projects():
            student_names = []  # type: List[str]
            supervisor_dissatisfactions = []  # type: List[float]
            student_dissatisfactions = []  # type: List[float]
            for student in self.allocated_students(project):
                student_names.append(student.name)
                supervisor_dissatisfactions.append(
                    project.dissatisfaction(student)
                )
                student_dissatisfactions.append(
                    student.dissatisfaction(project)
                )
            ps.append([
                project.name,
                project.supervisor.name,
                ", ".join(student_names),
                ", ".join(str(x) for x in student_dissatisfactions),
                ", ".join(str(x) for x in supervisor_dissatisfactions),
            ])

        # ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
        # Popularity of projects
        # ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
        pp = wb.create_sheet(SheetNames.PROJECT_POPULARITY)
        pp.append([
            "Project",
            "Supervisor",
            "Total dissatisfaction score from all students",
            "Number of students expressing a preference",
            "Students expressing a preference",
            "Allocated student(s)",
        ])
        proj_to_unpop = {}  # type: Dict[Project, float]
        for project in self.problem.projects:
            unpopularity = 0
            for student in self.problem.students:
                unpopularity += student.dissatisfaction(project)
            proj_to_unpop[project] = unpopularity
        for project, unpopularity in sorted(proj_to_unpop.items(),
                                            key=operator.itemgetter(1, 0)):
            allocated_students = ", ".join(
                student.name
                for student in self.allocated_students(project)
            )
            student_prefs = {}  # type: Dict[Student, float]
            for student in self.problem.students:
                if student.preferences.actively_expressed_preference_for(
                        project):
                    student_prefs[student] = student.preferences.preference(
                        project)
            student_details = []  # type: List[str]
            for student, studpref in sorted(student_prefs.items(),
                                            key=operator.itemgetter(1, 0)):
                student_details.append(f"{student.name} ({studpref})")
            pp.append([
                project.name,
                project.supervisor.name,
                unpopularity,
                len(student_details),
                ", ".join(student_details),
                allocated_students,
            ])

        # ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
        # Software, settings, and summary information
        # ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
        zs = wb.create_sheet(SheetNames.INFORMATION)
        is_stable, instability_reason = self.stability()
        zs_rows = [
            ["SOFTWARE DETAILS"],
            [],
            ["Software", "pdn_project_allocation"],
            ["Version", VERSION],
            ["Version date", VERSION_DATE],
            ["Source code",
             "https://github.com/RudolfCardinal/pdn_project_allocation"],
            ["Author", "Rudolf Cardinal (rudolf@pobox.com)"],
            [],
            ["RUN INFORMATION"],
            [],
            ["Date/time", datetime.datetime.now()],
            ["Overall weight given to student preferences",
             1 - self.problem.config.supervisor_weight],
            ["Overall weight given to supervisor preferences",
             self.problem.config.supervisor_weight],
            ["Command-line parameters", cmdline_quote(sys.argv)],
            ["Config", str(self.problem.config)],
            [],
            ["SUMMARY STATISTICS"],
            [],
            ["Student dissatisfaction median",
             self.student_dissatisfaction_median()],
            ["Student dissatisfaction mean",
             self.student_dissatisfaction_mean()],
            ["Student dissatisfaction variance",
             self.student_dissatisfaction_variance()],
            ["Student dissatisfaction minimum",
             self.student_dissatisfaction_min()],
            ["Student dissatisfaction minimum",
             self.student_dissatisfaction_max()],
            [],
            ["Supervisor dissatisfaction (with each student) median",
             self.supervisor_dissatisfaction_median()],
            ["Supervisor dissatisfaction (with each student) mean",
             self.supervisor_dissatisfaction_mean()],
            ["Supervisor dissatisfaction (with each student) variance",
             self.supervisor_dissatisfaction_variance()],
            ["Supervisor dissatisfaction (with each student) minimum",
             self.supervisor_dissatisfaction_min()],
            ["Supervisor dissatisfaction (with each student) maximum",
             self.supervisor_dissatisfaction_max()],
            [],
            ["Stable marriages?", str(is_stable)],
            ["If unstable, reason:", instability_reason]
        ]
        for row in zs_rows:
            zs.append(row)

        # ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
        # Problem definition
        # ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
        self.problem.write_to_xlsx_workbook(wb)

        wb.save(filename)

    def write_data(self, filename: str) -> None:
        """
        Autodetects the file type from the extension and writes data to that
        file.
        """
        # File type?
        _, ext = os.path.splitext(filename)
        if ext == EXT_XLSX:
            self.write_xlsx(filename)
        else:
            raise ValueError(
                f"Don't know how to write file type {ext!r} for {filename!r}")

    def write_student_csv(self, filename: str) -> None:
        """
        Writes just the "per student" mapping to a CSV file, for comparisons
        (e.g. via ``meld``).
        """
        log.info(f"Writing student allocation data to: {filename}")
        with open(filename, "w") as file:
            writer = csv.writer(file)
            writer.writerow([
                "Student number",
                "Student name",
                "Project number",
                "Project name",
                "Student's rank of (dissatisfaction with) allocated project",
            ])
            for student, project in self._gen_student_project_pairs():
                writer.writerow([
                    student.number,
                    student.name,
                    project.number,
                    project.name,
                    student.dissatisfaction(project),
                ])


# =============================================================================
# Eligibility helpers
# =============================================================================

class Eligibility(object):
    """
    Simple wrapper around a map between students and projects.
    """

    def __init__(self,
                 students: List[Student],
                 projects: List[Project],
                 default_eligibility: bool = True,
                 allow_defunct_projects: bool = False) -> None:
        """
        Default constructor, which just sets default eligibility for everyone.

        Args:
            projects:
                All projects.
            students:
                All students.
            default_eligibility:
                Default value for "is student eligible for project"?
            allow_defunct_projects:
                Allow projects that permit no students?
        """
        self.students = sorted(students, key=lambda s: s.number)
        self.projects = sorted(projects, key=lambda p: p.number)
        self.eligibility = OrderedDict(
            (
                s,
                OrderedDict(
                    (p, default_eligibility)
                    for p in projects
                )
            )
            for s in students
        )
        self.allow_defunct_projects = allow_defunct_projects

    def __str__(self) -> str:
        """
        String representations.
        """
        if self.everyone_eligible_for_everything():
            return "All students eligible for all projects."
        lines = []  # type: List[str]
        for s, p_e in self.eligibility.items():
            projects_str = ", ".join(
                str(p)
                for p, e in p_e.items()
                if e
            )
            lines.append(f"{s}: eligible for {projects_str}")
        return "\n".join(lines)

    def assert_valid(self) -> None:
        """
        Perform internal checks, or raise an exception.
        """
        # 1. Every student has an eligible project.
        for s in self.students:
            assert any(self.is_eligible(s, p) for p in self.projects), (
                f"Error: student {s} is not eligible for any projects!"
            )
        # 2. Every project has an eligible student.
        for p in self.projects:
            if not any(self.is_eligible(s, p) for s in self.students):
                msg = f"Project {p} has no eligible students!"
                if self.allow_defunct_projects:
                    log.warning(msg)
                else:
                    raise AssertionError(
                        msg + " [If you meant this, set the "
                              "--allow_defunct_projects option.]")

    def set_eligibility(self,
                        student: Student,
                        project: Project,
                        eligible: bool):
        """
        Set eligibility for a specific student/project combination.

        Args:
            student: the student
            project: the project
            eligible: is the student eligible for the project?
        """
        self.eligibility[student][project] = eligible

    def is_eligible(self, student: Student, project: Project) -> bool:
        """
        Is the student eligible for the project?
        """
        return self.eligibility[student][project]

    def everyone_eligible_for_everything(self) -> bool:
        """
        Is this a simple problem in which everyone is eligible for everything?
        """
        return all(
            e
            for p_e in self.eligibility.values()
            for e in p_e.values()
        )


# =============================================================================
# Problem
# =============================================================================

class Problem(object):
    """
    Represents the problem (and solves it) -- projects (with their supervisor's
    preferences for students), students (with their preferences for projects),
    and eligibility (which students are allowed to do which project?).
    """
    def __init__(self,
                 supervisors: List[Supervisor],
                 projects: List[Project],
                 students: List[Student],
                 config: Config,
                 eligibility: Eligibility = None) -> None:
        """
        Args:
            supervisors:
                List of project supervisors.
            projects:
                List of projects (with supervisor preference information).
            students:
                List of students (with their project preferences).
            config:
                Master config object
            eligibility:
                Dictionary of the form eligibility[student][project] = bool.

        Note that the students and projects are put into a "deterministic
        random" order, i.e. deterministically sorted, then shuffled (but with a
        globally fixed random number generator seed). That ensures fairness and
        consistency.
        """
        self.supervisors = supervisors
        self.projects = projects
        self.students = students
        self.config = config
        self.eligibility = eligibility or Eligibility(students, projects)
        self.eligibility.assert_valid()
        # Fix the order:
        if not config.no_shuffle:
            self.students.sort()
            random.shuffle(self.students)
            self.projects.sort()
            random.shuffle(self.projects)

    # -------------------------------------------------------------------------
    # Representations
    # -------------------------------------------------------------------------

    def __str__(self) -> str:
        """
        We re-sort the output for display purposes.
        """
        supervisors = "\n".join(
            sv.description() for sv in self.sorted_supervisors())
        projects = "\n".join(p.description() for p in self.sorted_projects())
        students = "\n".join(s.description() for s in self.sorted_students())
        return (
            f"Problem:\n"
            f"\n"
            f"- Supervisors:\n\n{supervisors}\n"
            f"\n"
            f"- Projects:\n\n{projects}\n"
            f"\n"
            f"- Students:\n\n{students}\n"
            f"\n"
            f"- Eligibility:\n\n{self.eligibility}\n"
        )

    # -------------------------------------------------------------------------
    # Information
    # -------------------------------------------------------------------------

    def sorted_supervisors(self) -> List[Supervisor]:
        """
        Supervisors, sorted by number.
        """
        return sorted(self.supervisors, key=lambda sv: sv.number)

    def sorted_students(self) -> List[Student]:
        """
        Students, sorted by number.
        """
        return sorted(self.students, key=lambda s: s.number)

    def sorted_projects(self) -> List[Project]:
        """
        Projects, sorted by number.
        """
        return sorted(self.projects, key=lambda p: p.number)

    def n_supervisors(self) -> int:
        """
        Number of supervisors.
        """
        return len(self.supervisors)

    def n_students(self) -> int:
        """
        Number of students.
        """
        return len(self.students)

    def n_projects(self) -> int:
        """
        Number of projects.
        """
        return len(self.projects)

    def students_who_chose(self, project: Project) -> List[Student]:
        """
        All students who ranked this project in some way.
        """
        return [
            s
            for s in self.students
            if s.preferences.actively_expressed_preference_for(project)
        ]

    def gen_student_project_pairs_where_student_chose_project(self) \
            -> Generator[Tuple[Student, Project], None, None]:
        """
        Generate ``student, project`` tuples where the student expressed some
        interest in the project.
        """
        for s in self.students:
            for p in s.preferences.items_explicitly_ranked():
                yield s, p

    def is_student_interested(self,
                              student: Student,
                              project: Project) -> bool:
        """
        Is the student interested in this project?
        """
        return self.students[student].actively_expressed_preference_for(project)  # noqa

    def are_preferences_strict_over_relevant_combos(self) -> bool:
        """
        Are all preferences strict, across combinations that matter?
        """
        # Students should strictly order their projects:
        for s in self.students:
            if not s.preferences.is_strict_over_expressed_preferences():
                return False
        # Supervisors should strictly order the students who expressed an
        # interest in their projects:
        for p in self.projects:
            students = self.students_who_chose(p)
            if not p.supervisor_preferences.is_strict_over(students):
                return False
        return True

    def gen_better_projects(
            self,
            student: Student,
            project: Project) -> Generator[Project, None, None]:
        """
        Generates projects that this student prefers over the specified one
        (and for which they're eligible).
        """
        current_dissatisfaction = student.dissatisfaction(project)
        for p in self.projects:
            if not self.eligibility.is_eligible(student, p):
                continue
            new_dissatisfaction = student.dissatisfaction(p)
            if new_dissatisfaction < current_dissatisfaction:
                yield p

    def gen_better_students(
            self,
            project: Project,
            student: Student) -> Generator[Student, None, None]:
        """
        Generates students that this project prefers over the specified one
        (and for which they're eligible).
        """
        current_dissatisfaction = project.dissatisfaction(student)
        for s in self.students:
            if not self.eligibility.is_eligible(s, project):
                continue
            new_dissatisfaction = project.dissatisfaction(s)
            if new_dissatisfaction < current_dissatisfaction:
                yield s

    def gen_worse_students(
            self,
            project: Project,
            student: Student) -> Generator[Student, None, None]:
        """
        Generates students that this project prefers LESS THAN the specified
        one (and for which they're eligible).
        """
        current_dissatisfaction = project.dissatisfaction(student)
        for s in self.students:
            if not self.eligibility.is_eligible(s, project):
                continue
            new_dissatisfaction = project.dissatisfaction(s)
            if new_dissatisfaction > current_dissatisfaction:
                yield s

    # -------------------------------------------------------------------------
    # Read data
    # -------------------------------------------------------------------------

    @classmethod
    def read_data(cls, config: Config) -> "Problem":
        """
        Reads a file, autodetecting its format, and returning the
        :class:`Problem`.
        """
        # File type?
        _, ext = os.path.splitext(config.filename)
        if ext == EXT_XLSX:
            return cls.read_data_xlsx(config)
        else:
            raise ValueError(f"Don't know how to read file type {ext!r} "
                             f"for {config.filename!r}")

    # noinspection DuplicatedCode
    @classmethod
    def read_data_xlsx(cls, config: Config) -> "Problem":
        """
        Reads a :class:`Problem` from an Excel XLSX file.
        """
        log.info(f"Reading XLSX file: {config.filename}")
        wb = load_workbook(config.filename, read_only=True, keep_vba=False,
                           data_only=True,  keep_links=False)

        # ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
        # Supervisors
        # ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~

        log.info("... reading supervisors...")
        supervisors = []  # type: List[Supervisor]
        sv_name_to_supervisor = {}  # type: Dict[str, Supervisor]
        # This will raise an error if the named sheet does not exist:
        ws_supervisors = wb[SheetNames.SUPERVISORS]  # type: Worksheet
        expected_headings = [
            SheetHeadings.SUPERVISOR,
            SheetHeadings.MAX_NUMBER_OF_PROJECTS,
            SheetHeadings.MAX_NUMBER_OF_STUDENTS,
        ]
        obtained_headings = [c.value for c in ws_supervisors["A1:C1"][0]]
        assert obtained_headings == expected_headings, (
            f"Bad headings to worksheet {SheetNames.SUPERVISORS}; expected "
            f"{expected_headings!r}, got {obtained_headings!r}"
        )
        sv_rows = read_until_empty_row(ws_supervisors)
        for row_number, row in enumerate(sv_rows[1:], start=2):
            supervisor_number = row_number - 1
            supervisor_name = row[0]
            assert supervisor_name, (
                f"Missing supervisor name in {SheetNames.SUPERVISORS} "
                f"row {row_number}"
            )
            assert supervisor_name not in sv_name_to_supervisor, (
                f"Duplicate supervisor name in {SheetNames.SUPERVISORS} "
                f"row {row_number}: {supervisor_name!r}"
            )
            max_n_projects = row[1]
            assert max_n_projects is None or isinstance(max_n_projects, int), (
                f"Max max_n_projects in {SheetNames.SUPERVISORS} "
                f"row {row_number}; is {max_n_projects!r}"
            )
            max_n_students = row[2]
            assert max_n_students is None or isinstance(max_n_students, int), (
                f"Max max_n_students in {SheetNames.SUPERVISORS} "
                f"row {row_number}; is {max_n_students!r}"
            )
            new_supervisor = Supervisor(
                name=supervisor_name,
                number=supervisor_number,
                max_n_projects=max_n_projects,
                max_n_students=max_n_students
            )
            sv_name_to_supervisor[supervisor_name] = new_supervisor
            supervisors.append(new_supervisor)
        n_supervisors = len(supervisors)
        assert n_supervisors, "No supervisors defined!"
        log.info(f"Number of supervisors: {n_supervisors}")
        del expected_headings, obtained_headings
        del ws_supervisors, sv_rows, row_number, row, supervisor_number
        del supervisor_name, max_n_projects, max_n_students, new_supervisor

        # ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
        # Projects
        # ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~

        log.info("... reading projects...")
        projects = []  # type: List[Project]
        ws_projects = wb[SheetNames.PROJECTS]  # type: Worksheet
        project_names = []  # type: List[str]
        expected_headings = [
            SheetHeadings.PROJECT,
            SheetHeadings.MAX_NUMBER_OF_STUDENTS,
            SheetHeadings.SUPERVISOR,
        ]
        obtained_headings = [c.value for c in ws_projects["A1:C1"][0]]
        assert obtained_headings == expected_headings, (
            f"Bad headings to worksheet {SheetNames.PROJECTS}; expected "
            f"{expected_headings!r}, got {obtained_headings!r}"
        )
        # log.debug(f"Projects: max_row = {ws_projects.max_row}")
        p_rows = read_until_empty_row(ws_projects)
        for row_number, row in enumerate(p_rows[1:], start=2):
            project_number = row_number - 1
            project_name = row[0]
            assert project_name, (
                f"Missing project name in {SheetNames.PROJECTS} "
                f"row {row_number}"
            )
            assert project_name not in project_names, (
                f"Duplicate project name in {SheetNames.PROJECTS} "
                f"row {row_number}: {project_name!r}"
            )
            max_n_students = row[1]
            assert isinstance(max_n_students, int), (
                f"Bad max_n_students in {SheetNames.PROJECTS} "
                f"row {row_number}; is {max_n_students!r}"
            )
            supervisor_name = row[2]
            assert supervisor_name in sv_name_to_supervisor, (
                f"Unknown supervisor in {SheetNames.PROJECTS} "
                f"row {row_number}: {supervisor_name!r}"
            )
            project_names.append(project_name)
            projects.append(Project(
                name=project_name,
                number=project_number,
                supervisor=sv_name_to_supervisor[supervisor_name],
                max_n_students=max_n_students,
                allow_defunct_projects=config.allow_defunct_projects
            ))
        n_projects = len(projects)
        assert n_projects, "No projects defined!"
        log.info(f"Number of projects: {n_projects}")
        del expected_headings, obtained_headings
        del ws_projects, p_rows, row_number, row, project_number, project_name
        del max_n_students, supervisor_name, project_names

        # ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
        # Students with their preferences
        # ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~

        log.info("... reading students and their preferences...")
        students = []  # type: List[Student]
        student_names = []  # type: List[str]
        ws_students = wb[SheetNames.STUDENT_PREFERENCES]  # type: Worksheet  # noqa
        stp_rows = read_until_empty_row(ws_students)
        # Check project headings
        assert all(
            stp_rows[0][i + 1] == projects[i].name
            for i in range(n_projects)
        ), (
            f"First row of {SheetNames.STUDENT_PREFERENCES} sheet "
            f"must contain all project names in the same order as in the "
            f"{SheetNames.PROJECTS} sheet"
        )
        for row_number, row in enumerate(stp_rows[1:], start=2):
            student_number = row_number - 1
            assert len(row) == n_projects + 1, (
                f"In {SheetNames.STUDENT_PREFERENCES}, student on row "
                f"{row_number} has a preference row of the wrong "
                f"length (expected {n_projects + 1})."
            )
            student_name = row[0]
            assert student_name not in student_names, (
                f"Duplicate student name in {SheetNames.STUDENT_PREFERENCES} "
                f"row {row_number}: {student_name!r}"
            )
            student_preferences = OrderedDict()  # type: Dict[Project, int]
            for project_number, pref in enumerate(row[1:], start=1):
                if config.allow_student_preference_ties:
                    ok = pref is None or isinstance(pref, float)
                else:
                    ok = pref is None or isinstance(pref, int)
                assert ok, (
                    f"Bad preference for student {student_name} in "
                    f"{SheetNames.STUDENT_PREFERENCES} "
                    f"row {row_number}: {pref!r}"
                )
                project = projects[project_number - 1]
                student_preferences[project] = pref
            student_names.append(student_name)
            new_student = Student(
                name=student_name,
                number=student_number,
                preferences=student_preferences,
                n_projects=n_projects,
                allow_ties=config.allow_student_preference_ties,
                preference_power=config.preference_power,
            )
            students.append(new_student)
            # log.critical(new_student)
        del ws_students, stp_rows, row_number, row, student_name, student_names
        del student_preferences, project_number, pref, ok, project, new_student
        n_students = len(students)
        log.info(f"Number of students: {n_students}")
        assert n_students >= 1

        # ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
        # Supervisor preferences, per project
        # ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~

        log.info("... reading supervisor preferences...")
        ws_supervisorprefs = wb[SheetNames.SUPERVISOR_PREFERENCES]  # type: Worksheet  # noqa
        # Accessing cells by (row, column) index is ridiculously slow here, and
        # the time is spent in the internals of openpyxl; specifically, in
        # xml.etree.ElementTree.XMLParser.feed(). That's true even after
        # install lxml as recommended, and specifying the "simple read-only"
        # options. So, it is **much** faster to load all the values like this
        # and then operate on the copies (e.g. ~6 seconds becomes ~1 ms):
        svp_rows = read_until_empty_row(ws_supervisorprefs)
        # ... index as : svp_rows[row_zero_based][column_zero_based]

        # Check project headings
        assert len(svp_rows[0]) == 1 + n_projects, (
            f"First row of {SheetNames.SUPERVISOR_PREFERENCES} should have "
            f"{1 + n_projects} columns (one on the left for student names "
            f"plus {n_projects} columns for projects). Yours has "
            f"{len(svp_rows[0])}."
        )
        assert all(
            svp_rows[0][i + 1] == projects[i].name
            for i in range(n_projects)
        ), (
            f"First row of {SheetNames.SUPERVISOR_PREFERENCES} sheet "
            f"must contain all project names in the same order as in the "
            f"{SheetNames.PROJECTS} sheet"
        )
        # Check student names
        assert len(svp_rows) == 1 + n_students, (
            f"Sheet {SheetNames.SUPERVISOR_PREFERENCES} should have "
            f"{1 + n_students} rows (one header row plus {n_students} "
            f"rows for students). Yours has {len(svp_rows)}."
        )
        _sn_from_sheet = [svp_rows[i + 1][0] for i in range(n_students)]
        _sn_from_students = [students[i].name for i in range(n_students)]
        assert _sn_from_sheet == _sn_from_students, (
            f"First column of {SheetNames.SUPERVISOR_PREFERENCES} sheet "
            f"must contain all student names in the same order as in the "
            f"{SheetNames.STUDENT_PREFERENCES} sheet. Mismatch is: "
            f"{mismatch(_sn_from_sheet, _sn_from_students)}"
        )
        # Read preferences
        for pcol, project in enumerate(projects, start=2):
            supervisor_prefs = OrderedDict()  # type: Dict[Student, int]
            for srow, student in enumerate(students, start=2):
                pref_value = svp_rows[srow - 1][pcol - 1]
                try:
                    pref = pref_value or None
                except (ValueError, TypeError):
                    raise ValueError(
                        f"Bad preference at row={srow}, col={pcol} in "
                        f"{SheetNames.SUPERVISOR_PREFERENCES}")
                supervisor_prefs[student] = pref
            project.set_supervisor_preferences(
                n_students=n_students,
                preferences=supervisor_prefs,
                allow_ties=config.allow_supervisor_preference_ties,
                preference_power=config.preference_power,
            )
        del ws_supervisorprefs, svp_rows, _sn_from_sheet, _sn_from_students
        del pcol, project, supervisor_prefs, srow, student, pref_value, pref

        # ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
        # Eligibility
        # ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~

        log.info("... reading eligibility...")
        eligibility = Eligibility(
            students=students,
            projects=projects,
            default_eligibility=True,
            allow_defunct_projects=config.allow_defunct_projects
        )
        if SheetNames.ELIGIBILITY in wb:
            ws_eligibility = wb[SheetNames.ELIGIBILITY]
            el_rows = read_until_empty_row(ws_eligibility)
            # ... index as : el_rows[row_zero_based][column_zero_based]
            # Check project headings
            assert all(
                el_rows[0][i + 1] == projects[i].name
                for i in range(n_projects)
            ), (
                f"First row of {SheetNames.ELIGIBILITY} sheet "
                f"must contain all project names in the same order as in the "
                f"{SheetNames.PROJECTS} sheet"
            )
            # Check student names
            _sn_from_sheet = [el_rows[i + 1][0] for i in range(n_students)]
            _sn_from_students = [students[i].name for i in range(n_students)]  # noqa
            assert _sn_from_sheet == _sn_from_students, (
                f"First column of {SheetNames.ELIGIBILITY} sheet "
                f"must contain all student names in the same order as in the "
                f"{SheetNames.STUDENT_PREFERENCES} sheet. Mismatch is: "
                f"{mismatch(_sn_from_sheet, _sn_from_students)}"
            )
            # Read eligibility
            for pcol, project in enumerate(projects, start=2):
                for srow, student in enumerate(students, start=2):
                    eligibility_val = el_rows[srow - 1][pcol - 1]
                    if eligibility_val in TRUE_VALUES:
                        eligible = True
                    elif eligibility_val in FALSE_VALUES:
                        eligible = False
                    elif (eligibility_val in MISSING_VALUES and
                            config.missing_eligibility is not None):
                        eligible = config.missing_eligibility
                    else:
                        raise ValueError(
                            f"Eligibility value {eligibility_val!r} "
                            f"(at row {srow}, column {pcol}) is "
                            f"invalid; use one of {TRUE_VALUES} "
                            f"for 'eligible', or one of {FALSE_VALUES} "
                            f"for 'ineligible'. The meaning of "
                            f"{MISSING_VALUES} is configurable."
                        )
                    if (config.student_must_have_choice and
                            not student.explicitly_ranked_project(project)):
                        eligible = False
                    eligibility.set_eligibility(student, project, eligible)
            del ws_eligibility, el_rows, _sn_from_sheet, _sn_from_students
            del pcol, project, eligibility_val, eligible

        # ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
        # Create and return the Problem object
        # ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~

        log.info("... finished reading")
        return Problem(supervisors=supervisors,
                       projects=projects,
                       students=students,
                       eligibility=eligibility,
                       config=config)

    # -------------------------------------------------------------------------
    # Save data
    # -------------------------------------------------------------------------

    # noinspection DuplicatedCode
    def write_to_xlsx_workbook(self, wb: Workbook) -> None:
        """
        Writes the problem data to a spreadsheet (so it can be saved alongside
        the solution).

        Args:
            wb:
                A :class:`openpyxl.workbook.workbook.Workbook` to which to
                write.
        """
        sorted_projects = self.sorted_projects()
        sorted_students = self.sorted_students()

        # ---------------------------------------------------------------------
        # Supervisors
        # ---------------------------------------------------------------------

        supervisor_sheet = wb.create_sheet(SheetNames.SUPERVISORS)
        supervisor_sheet.append([
            SheetHeadings.SUPERVISOR,
            SheetHeadings.MAX_NUMBER_OF_PROJECTS,
            SheetHeadings.MAX_NUMBER_OF_STUDENTS
        ])
        for sv in self.sorted_supervisors():
            supervisor_sheet.append([
                sv.name,
                sv.max_n_projects,
                sv.max_n_students
            ])

        # ---------------------------------------------------------------------
        # Projects
        # ---------------------------------------------------------------------

        project_sheet = wb.create_sheet(SheetNames.PROJECTS)
        project_sheet.append([
            SheetHeadings.PROJECT,
            SheetHeadings.MAX_NUMBER_OF_STUDENTS
        ])
        for p in sorted_projects:
            project_sheet.append([
                p.name,
                p.max_n_students,
                p.supervisor.name
            ])

        # ---------------------------------------------------------------------
        # Students
        # ---------------------------------------------------------------------

        student_sheet = wb.create_sheet(SheetNames.STUDENT_PREFERENCES)
        student_sheet.append(
            [""] + [p.name for p in sorted_projects]
        )
        for s in sorted_students:
            # noinspection PyTypeChecker
            student_sheet.append(
                [s.name] + [s.preferences.raw_preference(p)
                            for p in sorted_projects]
            )

        # ---------------------------------------------------------------------
        # Supervisor preferences
        # ---------------------------------------------------------------------

        supervisor_sheet = wb.create_sheet(SheetNames.SUPERVISOR_PREFERENCES)
        supervisor_sheet.append(
            [""] + [p.name for p in sorted_projects]
        )
        for s in sorted_students:
            # noinspection PyTypeChecker
            supervisor_sheet.append(
                [s.name] + [p.supervisor_preferences.raw_preference(s)
                            for p in sorted_projects]
            )

        # ---------------------------------------------------------------------
        # Eligibility
        # ---------------------------------------------------------------------

        eligibility_sheet = wb.create_sheet(SheetNames.ELIGIBILITY)
        eligibility_sheet.append(
            [""] + [p.name for p in sorted_projects]
        )
        for s in sorted_students:
            # noinspection PyTypeChecker
            eligibility_sheet.append(
                [s.name] + [int(self.eligibility.is_eligible(s, p))
                            for p in sorted_projects]
            )

    # -------------------------------------------------------------------------
    # Solver entry point
    # -------------------------------------------------------------------------

    def best_solution(self) -> Optional[Solution]:
        """
        Return the best solution.
        """
        method = self.config.optimize_method
        if method == OptimizeMethod.MINIMIZE_DISSATISFACTION:
            return self.best_solution_mip(enforce_stability=False)
        elif method == OptimizeMethod.MINIMIZE_DISSATISFACTION_STABLE_AB1996:
            return self.best_solution_mip(
                enforce_stability=True, stability_ab1996=True)
        elif method == OptimizeMethod.MINIMIZE_DISSATISFACTION_STABLE_CUSTOM:
            return self.best_solution_mip(
                enforce_stability=True, stability_ab1996=False)
        elif method == OptimizeMethod.MINIMIZE_DISSATISFACTION_STABLE:
            return (
                self.best_solution_mip(
                    enforce_stability=True, stability_ab1996=True) or
                self.best_solution_mip(
                    enforce_stability=True, stability_ab1996=False)
            )
        elif method == OptimizeMethod.MINIMIZE_DISSATISFACTION_STABLE_FALLBACK:
            solution = (
                self.best_solution_mip(
                    enforce_stability=True, stability_ab1996=True) or
                self.best_solution_mip(
                    enforce_stability=True, stability_ab1996=False)
            )
            if solution:
                return solution
            log.warning("Stable solution not found. Falling back to "
                        "overall best (permitting instability).")
            return self.best_solution_mip(enforce_stability=False)
        elif method == OptimizeMethod.ABRAHAM_STUDENT:
            return self.best_solution_abraham(optimal="student")
        elif method == OptimizeMethod.ABRAHAM_SUPERVISOR:
            return self.best_solution_abraham(optimal="supervisor")
        else:
            raise AssertionError(f"Unknown optimization method: {method!r}")

    # -------------------------------------------------------------------------
    # Solve via MIP
    # -------------------------------------------------------------------------

    def best_solution_mip(
            self,
            enforce_stability: bool = False,
            stability_ab1996: bool = False) -> Optional[Solution]:
        """
        Return the best solution by optimizing with the MIP package.
        This is extremely impressive.
        See https://python-mip.readthedocs.io/.

        Args:
            enforce_stability:
                Ensure only stable "marriages" are produced (or fail entirely)?
            stability_ab1996:
                For ``enforce_stability``: use the stability constraint that is
                equation 4 of Abeledo & Blum (1996,
                https://doi.org/10.1016/0024-3795(95)00052-6).
        """
        # ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
        # Basic setup
        # ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
        supervisor_weight = self.config.supervisor_weight
        assert 0 <= supervisor_weight <= 1
        student_weight = 1 - supervisor_weight
        log.info(
            f"MIP approach: student_weight={student_weight}, "
            f"supervisor_weight={supervisor_weight}, "
            f"enforce_stability={enforce_stability}, "
            f"stability_ab1996={stability_ab1996}")
        n_students = len(self.students)
        n_projects = len(self.projects)
        using_max_projects_per_supervisor = any(
            supervisor.max_n_projects is not None
            for supervisor in self.supervisors
        )

        # ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
        # Eligibility map
        # ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
        eligible = [
            [
                self.eligibility.is_eligible(student, project)
                for p, project in enumerate(self.projects)  # second index
            ]
            for s, student in enumerate(self.students)  # first index
        ]  # indexed s, p

        # ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
        # Dissatisfaction scores for each project/student combination
        # ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
        # CAUTION: get indexes the right way round!
        student_dissatisfaction_with_project = [
            [
                self.students[s].exponentiated_dissatisfaction(self.projects[p])  # noqa
                for p in range(n_projects)  # second index
            ]
            for s in range(n_students)  # first index
        ]  # indexed s, p
        project_dissatisfaction_with_student = [
            [
                self.projects[p].exponentiated_dissatisfaction(self.students[s])  # noqa
                for p in range(n_projects)  # second index
            ]
            for s in range(n_students)  # first index
        ]  # indexed s, p
        weighted_dissatisfaction = [
            [
                (
                    student_weight *
                    student_dissatisfaction_with_project[s][p] +
                    supervisor_weight *
                    project_dissatisfaction_with_student[s][p]
                )
                for p in range(n_projects)  # second index
            ]
            for s in range(n_students)  # first index
        ]  # indexed s, p

        # ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
        # Model
        # ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
        m = Model("Student project allocation")
        # Binary variables to optimize, each linking a student to a project
        # CAUTION: get indexes the right way round!
        x = [
            [
                (
                    m.add_var(f"x[s={s},p={p}]", var_type=BINARY)
                    if eligible[s][p] else None
                )
                for p in range(n_projects)  # second index
            ]
            for s in range(n_students)  # first index
        ]  # indexed s, p
        if using_max_projects_per_supervisor:
            # See below for explanation.
            project_in_use = [
                m.add_var(f"project_in_use[p={p}]", var_type=BINARY)
                if self.projects[p].supervisor.max_n_projects is not None
                else None  # don't bother for supervisors that don't care
                for p in range(n_projects)
            ]  # indexed: p
        else:
            project_in_use = []  # type: List[Var]

        # ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
        # Objective: happy students/supervisors
        # ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
        m.objective = minimize(xsum(
            x[s][p] * weighted_dissatisfaction[s][p]
            for p in range(n_projects)
            for s in range(n_students)
            if eligible[s][p]
        ))

        # ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
        # Constraint: For each student, exactly one project.
        # ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
        for s in range(n_students):
            m += (
                xsum(x[s][p]
                     for p in range(n_projects)
                     if eligible[s][p]) == 1,
                f"student_{s}_one_project"
            )
            # Using a Special Ordered Set here doesn't materially speed things
            # up (maybe very slightly). I'm not entirely sure what the "weight"
            # parameter should be. Always 1? Or consecutive?
            # - https://docs.python-mip.com/en/latest/examples.html#exsos
            #   ... not terribly clear, but does use non-sequential order (of
            #   possible plants in a region) as weights.
            # - http://lpsolve.sourceforge.net/5.5/SOS.htm
            # - https://en.wikipedia.org/wiki/Special_ordered_set
            #   ... the benefit is for speed.
            # - https://www.tu-chemnitz.de/mathematik/discrete/manuals/cplex/doc/pdf/cplex81userman.pdf  # noqa
            #   ... gives an example (p244) using ordered warehouse size as the
            #   weights.
            # So:
            m.add_sos(
                sos=[
                    (x[s][p], p)  # p (the non-sequential order) is the weight
                    for p in range(n_projects)
                    if eligible[s][p]
                ],
                sos_type=1  # Type 1: only one variable can receive value 1.
            )
        del s

        # ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
        # Constraint: For each project, up to the maximum number of students.
        # ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
        for p, project in enumerate(self.projects):
            m += (
                xsum(x[s][p]
                     for s in range(n_students)
                     if eligible[s][p]) <= project.max_n_students,
                f"project_{p}_max_{project.max_n_students}_students"
            )
        del p, project

        # ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
        # Constraint: Maximum number of projects per supervisor
        # ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
        # This one is hard.
        # General advice on formulating these problems:
        # - http://people.brunel.ac.uk/~mastjjb/jeb/or/moreip.html
        # - https://pubsonline.informs.org/doi/pdf/10.1287/ited.2017.0177
        #   ^^^
        #   THIS ONE! Stevens & Palocsay (2017). Excellent.
        # Somewhat related problems:
        # - http://yetanothermathprogrammingconsultant.blogspot.com/2018/04/a-difficult-mip-construct-counting.html  # noqa
        # - https://math.stackexchange.com/questions/2732897/linear-integer-programming-count-consecutive-ones  # noqa
        if using_max_projects_per_supervisor:
            # - We work out whether each project is allocated, using SEPARATE
            #   (BINARY) VARIABLES.
            # - The rule for project p is:
            #       "If any student is allocated to p, then project_in_use[p]."
            # - That is:
            #       "SOME student allocated -> ALL that project in use."
            # - By the Decomposition rule, that translates to:
            #       student 1 allocated to p -> project p is in use
            #       student 2 allocated to p -> project p is in use
            #       ...
            # - By the Translation rule, each one can be represented by
            #       student_1_allocated_to_p <= project_p_in_use
            # - Converting that to a form with constants on the right,
            #       student_1_allocated_to_p - project_p_in_use <= 0

            for sv, supervisor in enumerate(self.supervisors):
                if supervisor.max_n_projects is not None:
                    # 1. Define whether relevant projects are in use.
                    for p in range(n_projects):
                        if not self.projects[p].is_supervised_by(supervisor):
                            continue
                        for s in range(n_students):
                            m += (
                                x[s][p] - project_in_use[p] <= 0,
                                f"project_{p}_in_use_by_student_{s}"
                            )
                    # 2. Constrain the number of projects for the supervisor.
                    m += (
                        xsum(
                            project_in_use[p]
                            for p in range(n_projects)
                            if self.projects[p].is_supervised_by(supervisor)
                        ) <= supervisor.max_n_projects,
                        f"supervisor_{sv}_max_{supervisor.max_n_projects}_projects"  # noqa
                    )
            del sv, supervisor

        # ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
        # Constraint: Maximum number of students per supervisor
        # ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
        for sv, supervisor in enumerate(self.supervisors):
            if supervisor.max_n_students is not None:
                m += (
                    xsum(
                        # "All students allocated to projects supervised by this
                        # supervisor."
                        x[s][p]
                        for s in range(n_students)
                        for p in range(n_projects)
                        if (
                            self.projects[p].is_supervised_by(supervisor) and
                            eligible[s][p]  # don't consider impossible pairings
                        )
                    ) <= supervisor.max_n_students,
                    f"supervisor_{sv}_max_{supervisor.max_n_students}_students"
                )
        del sv, supervisor

        # ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
        # Constraint: Only stable "marriages"?
        # ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
        if enforce_stability and stability_ab1996:
            # +++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
            # Stability via Abeledo & Blum 1996, assuming strict preferences
            # +++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
            log.info("Trying for stability via Abeledo & Blum 1996, which "
                     "assumes strict preferences.")
            if not self.are_preferences_strict_over_relevant_combos():
                log.error(
                    "Stability constraints of Abeledo & Blum (1996) require "
                    "strict preferences, but preferences are not strict (of "
                    "students for their projects, and of projects/supervisors "
                    "for all students who picked them). Failing.")
                return None
            # Equation 4 of Abeledo & Blum (1996), as above: the stability
            # constraint. We'll use their notation for clarity.
            # When they say a >{x} b, they mean "x prefers a to b".
            # Similarly, "a <{x} b" means "x prefers b to a".
            for u, v in product(range(n_students), range(n_projects)):
                # We'll say the student is "u" and the project is "v"
                if not eligible[u][v]:
                    continue
                student_dis = student_dissatisfaction_with_project[u][v]
                project_dis = project_dissatisfaction_with_student[u][v]
                other_project_vars = []  # type: List[Var]
                other_student_vars = []  # type: List[Var]
                for i in [_ for _ in range(n_projects) if _ != v]:  # "i"
                    if not eligible[u][i]:
                        continue
                    if (student_dissatisfaction_with_project[u][i] <
                            student_dis):
                        # Student "u" prefers project "i" to project "v";
                        # that is, i >{u} v.
                        other_project_vars.append(x[u][i])
                for j in [_ for _ in range(n_students) if _ != u]:  # "j"
                    if not eligible[j][v]:
                        continue
                    if (project_dissatisfaction_with_student[j][v] <
                            project_dis):
                        # Project "v" prefers student "j" to student "u";
                        # that is, j >{v} u.
                        other_student_vars.append(x[j][v])
                        # I'm pretty sure they must mean x{j,v} not x{v,j},
                        # since the variable x is always suffixed
                        # {u-type-thing, v-type-thing}, e.g. page 323.
                vars_to_sum = (
                    other_project_vars +  # sum{for i >{u} v}{x{u,i}}
                    other_student_vars +  # sum{for j >{v} u}{x{j,v}}
                    [x[u][v]]  # "x{u,v}"
                )
                stability_constraint = xsum(vars_to_sum) >= 1  # Eq. 4.
                log.debug(f"Adding stability constraint: "
                          f"{stability_constraint}")
                m += stability_constraint, f"stability_s{u}_p{v}"
            # What's the logic here?
            # Lemma 3.1, which includes equation 4, is from ref. [2], which is
            # Abeledo & Rothblum (1994,
            # https://doi.org/10.1016/0166-218X(94)90130-9).
            # In that work, it's Theorem 3.1, Equation 7, p6. The logic is:
            # - If there is no better match for either the student or the
            #   project, then the first two components vanish, x{u,v} is the
            #   best match, and must be 1. [RNC CAVEAT: THAT REQUIRES STRICT
            #   ORDERING, which is one of the assumptions. So we need to deal
            #   with that.] So this seems to be saying "if there's no better
            #   match, pick it".
            # - Combined with the other inequalities, which say things like
            #   "everyone must be assigned" and "not too many students per
            #   project", forcing us to pick the best excludes picking anything
            #   that isn't the best.
            # - [In passing, note that we have a "bipartite" situation
            #   (projects are distinct from students; defined on p3).]
            del u, v
            del vars_to_sum, stability_constraint

        elif enforce_stability and not stability_ab1996:
            # +++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
            # Stability via a custom approach allowing non-strict preferences
            # +++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
            # Can we develop an equivalent when there might be indifference?
            # We want to say simply "if there's a better marriage, don't pick
            # this one".
            log.info("Trying for stability via a custom method, which "
                     "does not assume strict preferences. (Can be slow.)")
            stability_constraints = set()  # type: Set[LinExpr]
            stability_constraint_tuples = []  # type: List[Tuple[LinExpr, str]]
            for s_idx, p_idx in product(range(n_students), range(n_projects)):
                if not eligible[s_idx][p_idx]:
                    continue
                s = self.students[s_idx]
                p = self.projects[p_idx]
                # So, for every eligible student/project combination...
                for other_p in self.gen_better_projects(s, p):
                    # other_p: "Other projects that s prefers to p."
                    other_p_idx = self.projects.index(other_p)
                    for other_s in self.gen_worse_students(other_p, s):
                        # other_s: "Other students that other_p would
                        # reject in favour of s."
                        other_s_idx = self.students.index(other_s)
                        constraint = (
                            # "Do not assign s to p and simultaneously
                            # assign other_s to other_p (because s and
                            # other_p would rather pair up with each
                            # other)." That is, s and other_p represent a
                            # blocking pair for a solution that includes a
                            # match between s and p and also between
                            # other_s and other_p.
                            x[s_idx][p_idx] + x[other_s_idx][other_p_idx] <= 1
                            # You can't multiply these variables, but you
                            # can add them.
                        )
                        if constraint not in stability_constraints:
                            # We use a set because otherwise we may add the
                            # same thing several times.
                            log.debug(
                                f"Adding stability constraint: {constraint}, "
                                f"for s={s}, p={p}, "
                                f"other_s={other_s}, other_p={other_p}"
                            )
                            stability_constraints.add(constraint)
                            stability_constraint_tuples.append((
                                constraint,
                                f"stability_s{s_idx}_p{p_idx}_"
                                f"other_s{other_s_idx}_other_p{other_p_idx}"
                            ))
                del s, p
            log.info(f"Adding {len(stability_constraints)} unique "
                     f"stability constraints")
            for stability_constraint_tuple in stability_constraint_tuples:
                m += stability_constraint_tuple
            del s_idx, p_idx
            del stability_constraints, stability_constraint_tuples

        # ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
        # Debug?
        # ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
        # import pdb; pdb.set_trace()
        if self.config.debug_model:
            report_on_model(m)

        # ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
        # Optimize
        # ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
        m.optimize(max_seconds=self.config.max_time_s)

        # ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
        # Debug?
        # ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
        if self.config.debug_model:
            report_on_model(m, solution_only=True)

        # ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
        # Extract results
        # ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
        if not m.num_solutions:
            return None
        project_indexes = [
            next(p for p in range(n_projects)
                 if eligible[s][p] and x[s][p].x >= ALMOST_ONE)
            # ... note that the value of a solved variable is var.x
            # If those two expressions are not the same, there's a bug.
            for s in range(n_students)
        ]
        solution = self._make_solution(project_indexes)
        if enforce_stability:
            assert solution.is_stable()
        return solution

    def _make_solution(self,
                       project_indexes: Sequence[int],
                       validate: bool = True) -> Solution:
        """
        Creates a solution from project index numbers.

        Args:
            project_indexes:
                Indexes (zero-based) of project numbers, one per student,
                in the order of ``self.students``.
            validate:
                validate input? For debugging only.
        """
        if validate:
            n_students = len(self.students)
            assert len(project_indexes) == n_students, (
                "Number of project indices does not match number of students"
            )
        allocation = {}  # type: Dict[Student, Project]
        for student_idx, project_idx in enumerate(project_indexes):
            allocation[self.students[student_idx]] = self.projects[project_idx]
        return Solution(problem=self, allocation=allocation)

    # -------------------------------------------------------------------------
    # Solve via Abraham-Irving-Manlove 2007
    # -------------------------------------------------------------------------

    def best_solution_abraham(self,
                              optimal: str = "student") -> Optional[Solution]:
        """
        Optimize via the Abraham-Irving-Manlove 2007 algorithm, optimally for
        students.

        For the Gale-Shapley algorithm, see

        - https://en.wikipedia.org/wiki/Gale%E2%80%93Shapley_algorithm
        - https://www.nrmp.org/nobel-prize/

        Others' work on Gale-Shapley:

        - https://towardsdatascience.com/gale-shapley-algorithm-simply-explained-caa344e643c2
        - https://gist.github.com/joyrexus/9967709 (a good one)
        - https://github.com/Vishal-Kancharla/Gale-Shapley-Algorithm
        - https://rosettacode.org/wiki/Stable_marriage_problem

        For Abraham-Irving-Manlove:

        - https://matching.readthedocs.io/
        """  # noqa
        assert optimal in ["student", "supervisor"]

        # ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
        # Set up the problem: (1) create objects
        # ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~

        mg_students = []  # type: List[MGStudent]
        student_to_mg_student = {}  # type: Dict[Student, MGStudent]
        mg_student_to_student = {}  # type: Dict[MGStudent, Student]
        for s in self.students:
            mg_student = MGStudent(name=s.name)
            mg_students.append(mg_student)
            student_to_mg_student[s] = mg_student
            mg_student_to_student[mg_student] = s

        mg_supervisors = []  # type: List[MGSupervisor]
        mg_projects = []  # type: List[MGProject]
        project_to_mg_supervisor = {}  # type: Dict[Project, MGSupervisor]
        mg_supervisor_to_project = {}  # type: Dict[MGSupervisor, Project]
        project_to_mg_project = {}  # type: Dict[Project, MGProject]
        mg_project_to_project = {}  # type: Dict[MGProject, Project]
        for p in self.projects:
            mg_supervisor = MGSupervisor(
                name=f"Supervisor of {p.name}",
                capacity=p.max_n_students
            )
            mg_supervisors.append(mg_supervisor)
            project_to_mg_supervisor[p] = mg_supervisor
            mg_supervisor_to_project[mg_supervisor] = p

            mg_project = MGProject(
                name=p.name,
                capacity=p.max_n_students
            )
            mg_project.set_supervisor(mg_supervisor)
            mg_projects.append(mg_project)
            project_to_mg_project[p] = mg_project
            mg_project_to_project[mg_project] = p

        # ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
        # Set up the problem: (2) define preferences
        # ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~

        # Student preferences. Must set these first.
        for s in self.students:
            mg_student = student_to_mg_student[s]
            preferred_projects = [
                project_to_mg_project[p]
                for p in s.projects_in_descending_order([
                    # Only the projects that the student has ranked
                    # explicitly...
                    px
                    for px in s.preferences.items_explicitly_ranked()
                    # ... and that the student is eligible for.
                    if self.eligibility.is_eligible(s, px)
                ])
            ]
            log.debug(f"For student {mg_student}, "
                      f"setting preferences: {preferred_projects}")
            mg_student.set_prefs(preferred_projects)

        # Supervisor/project preferences. (These are assigned to supervisors.)
        for p in self.projects:
            mg_supervisor = project_to_mg_supervisor[p]
            preferred_students = [
                student_to_mg_student[s]
                for s in p.students_in_descending_order([
                    # Only the students that explicitly chose this project...
                    sx
                    for sx in self.students_who_chose(p)
                    # ... and are eligible for it:
                    if self.eligibility.is_eligible(sx, p)
                ])
            ]
            log.debug(f"For supervisor {mg_supervisor}, "
                      f"setting preferences: {preferred_students}")
            mg_supervisor.set_prefs(preferred_students)

        # log.critical(f"Supervisors: {mg_supervisors}")
        # log.critical(f"Projects: {mg_projects}")
        # log.critical(f"Students: {mg_students}")

        # ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
        # Solve
        # ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
        game = MGStudentAllocation(
            mg_students, mg_projects, mg_supervisors)
        matching = game.solve(optimal=optimal)
        if matching is None:  # no solution found
            return None
        assert game.check_validity()
        assert game.check_stability()
        # log.critical(repr(matching))

        # ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
        # Translate back to our Solution class
        # ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
        allocation = {}  # type: Dict[Student, Project]
        for mg_project_copy, mg_student_copies in matching.items():
            # We cannot do a lookup. It has done a deepcopy. We have to match
            # by name.
            mg_project = next(mgp for mgp in mg_projects
                              if mgp.name == mg_project_copy.name)
            project = mg_project_to_project[mg_project]
            for mg_student_copy in mg_student_copies:
                # Ditto... this is a bit silly...
                mg_student = next(mgs for mgs in mg_students
                                  if mgs.name == mg_student_copy.name)
                student = mg_student_to_student[mg_student]
                allocation[student] = project
        # Create the solution
        solution = Solution(problem=self, allocation=allocation)
        # Sanity-check the solution
        unallocated_students = [
            s for s in self.students if s not in allocation
        ]
        if unallocated_students:
            log.critical(solution)
            log.critical(f"Failed: unallocated students: "
                         f"{unallocated_students}")
            return None
        return solution


# =============================================================================
# main
# =============================================================================

def main() -> None:
    """
    Command-line entry point.
    """
    # noinspection PyTypeChecker
    parser = argparse.ArgumentParser(
        formatter_class=RawDescriptionArgumentDefaultsHelpFormatter,
        description=f"""
Allocate students to projects, maximizing some version of happiness.

The input spreadsheet should have the following format (in each case, the
first row is the title row):

    Sheet name:
        {SheetNames.SUPERVISORS}
    Description:
        List of supervisors (one per row) and their project/student capacity.
        Capacity values can be left blank for "no maximum".
    Format:
        {SheetHeadings.SUPERVISOR}      {SheetHeadings.MAX_NUMBER_OF_PROJECTS}  {SheetHeadings.MAX_NUMBER_OF_STUDENTS}
        Dr Smith        3                       5
        Dr Jones
        Dr Lucas                                2
        ...             ...                     ...
        
    Sheet name:
        {SheetNames.PROJECTS}
    Description:
        List of projects (one per row), their student capacity, and their
        supervisor.
    Format:
        {SheetHeadings.PROJECT}         {SheetHeadings.MAX_NUMBER_OF_STUDENTS}  {SheetHeadings.SUPERVISOR}
        Project One     1                       Dr Jones
        Project Two     1                       Dr Jones
        Project Three   2                       Dr Smith
        ...             ...                     ...
        
    Sheet name:
        {SheetNames.STUDENT_PREFERENCES}
    Description:
        List of students (one per row) and their rank preferences (1 = top, 2 =
        next, etc.) for projects (one per column).
    Format:
        <ignored>       Project One     Project Two     Project Three   ...
        Miss Smith      1               2                               ...
        Mr Jones        2               1               3               ...
        ...             ...             ...             ...             ...
    
    Sheet name:
        {SheetNames.ELIGIBILITY}
    Description:
        OPTIONAL sheet, showing which students are eligible for which
        projects. If absent, all students are eligible for all projects.
        Use {TRUE_VALUES} for "eligible".
        Use {FALSE_VALUES} for "ineligible".
        Use --missing_eligibility to control the handling of empty cells.
    Format:
        <ignored>       Project One     Project Two     Project Three   ...
        Miss Smith      1               1               1               ...
        Mr Jones        1               0               1               ...
        ...             ...             ...             ...             ...
    
    Sheet name:
        {SheetNames.SUPERVISOR_PREFERENCES}
    Description:
        List of projects (one per column) and their supervisor's rank
        preferences (1 = top, 2 = next, etc.) for students (one per row).
    Format:
        <ignored>       Project One     Project Two     Project Three   ...
        Miss Smith      1               1                               ...
        Mr Jones        2               2                               ...
        ...             ...             ...

"""  # noqa
    )
    parser.add_argument(
        "--verbose", action="store_true",
        help="Be verbose"
    )

    file_group = parser.add_argument_group("Files")
    file_group.add_argument(
        "filename", type=str,
        help="Spreadsheet filename to read. "
             "Input file types supported: " + str(INPUT_TYPES_SUPPORTED)
    )
    file_group.add_argument(
        "--output", type=str,
        help="Optional filename to write output to. "
             "Output types supported: " + str(OUTPUT_TYPES_SUPPORTED)
    )
    file_group.add_argument(
        "--output_student_csv", type=str,
        help="Optional filename to write student CSV output to."
    )

    data_group = parser.add_argument_group("Data")
    data_group.add_argument(
        "--allow_student_preference_ties", action="store_true",
        help="Allow students to express tied preferences "
             "(e.g. 2.5 for joint second/third place)?"
    )
    data_group.add_argument(
        "--allow_supervisor_preference_ties", action="store_true",
        help="Allow supervisors to express tied preferences "
             "(e.g. 2.5 for joint second/third place)?"
    )
    data_group.add_argument(
        "--missing_eligibility", type=bool, default=None,
        help="If an eligibility cell is blank, treat it as eligible (use "
             "'True') or ineligible (use 'False')? Default, of None, means "
             "empty cells are invalid."
    )
    data_group.add_argument(
        "--allow_defunct_projects", action="store_true",
        help="Allow projects that say that all students are ineligible (e.g. "
             "because they've been pre-allocated by different process)?"
    )

    method_group = parser.add_argument_group("Method")
    method_group.add_argument(
        "--supervisor_weight", type=float, default=DEFAULT_SUPERVISOR_WEIGHT,
        help="Weight allocated to supervisor preferences (student preferences "
             "are weighted as [1 minus this])"
    )
    method_group.add_argument(
        "--preference_power", type=float, default=DEFAULT_PREFERENCE_POWER,
        help="Power (exponent) to raise preferences by."
    )
    method_group.add_argument(
        "--student_must_have_choice", action="store_true",
        help="Prevent students being allocated to projects they've not "
             "explicitly ranked?"
    )

    technical_group = parser.add_argument_group("Technicalities")
    technical_group.add_argument(
        "--maxtime", type=float, default=DEFAULT_MAX_SECONDS,
        help="Maximum time (in seconds) to run MIP optimizer for"
    )
    technical_group.add_argument(
        "--seed", type=int, default=None,
        help="Seed for random number generator. "
             "DO NOT USE FOR ACTUAL ALLOCATIONS; IT IS UNFAIR (because it "
             "tempts the operator to re-run with different seeds). "
             "FOR DEBUGGING USE ONLY."
    )
    technical_group.add_argument(
        "--no_shuffle", action="store_true",
        help="Don't shuffle anything. FOR DEBUGGING USE ONLY."
    )
    technical_group.add_argument(
        "--debug_model", action="store_true",
        help="Report the details of the MIP model before solving."
    )
    method_k, method_desc = keys_descriptions_from_enum(
        OptimizeMethod, keys_to_lower=True)
    method_group.add_argument(
        "--method", type=str, choices=method_k,
        default=DEFAULT_METHOD.name,
        help=f"Method of solving. -- {method_desc} --"
    )

    args = parser.parse_args()
    main_only_quicksetup_rootlogger(level=logging.DEBUG if args.verbose
                                    else logging.INFO)

    # Seed RNG
    if args.seed is not None:
        log.warning("You have specified --seed. FOR DEBUGGING USE ONLY: "
                    "THIS IS NOT FAIR FOR REAL ALLOCATIONS!")
        seed = args.seed
    else:
        seed = RNG_SEED
    random.seed(seed)

    # Go
    config = Config(
        allow_defunct_projects=args.allow_defunct_projects,
        allow_student_preference_ties=args.allow_student_preference_ties,
        allow_supervisor_preference_ties=args.allow_supervisor_preference_ties,
        cmd_args=vars(args),
        debug_model=args.debug_model,
        filename=args.filename,
        max_time_s=args.maxtime,
        missing_eligibility=args.missing_eligibility,
        no_shuffle=args.no_shuffle,
        optimize_method=OptimizeMethod[args.method],
        preference_power=args.preference_power,
        student_must_have_choice=args.student_must_have_choice,
        supervisor_weight=args.supervisor_weight,
    )
    log.info(f"Command: {cmdline_quote(sys.argv)}")
    log.info(f"Config: {config}")
    problem = Problem.read_data(config)
    if args.output:
        log.debug(problem)
    else:
        log.info(problem)
    solution = problem.best_solution()
    if solution:
        if args.output:
            log.debug(solution)
        else:
            log.info(solution)
        if args.output:
            solution.write_data(args.output)
        else:
            log.warning(
                "Output not saved. Specify the --output option for that.")
        if args.output_student_csv:
            solution.write_student_csv(args.output_student_csv)
        sys.exit(EXIT_SUCCESS)
    else:
        log.error("No solution found!")
        sys.exit(EXIT_FAILURE)


if __name__ == "__main__":
    try:
        main()
    except Exception as _top_level_exception:
        log.critical(str(_top_level_exception))
        log.critical(traceback.format_exc())
        sys.exit(EXIT_FAILURE)
