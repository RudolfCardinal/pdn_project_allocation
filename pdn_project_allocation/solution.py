#!/usr/bin/env python

"""
pdn_project_allocation/solution.py

===============================================================================

    Copyright (C) 2019 Rudolf Cardinal (rudolf@pobox.com).

    This file is part of pdn_project_allocation.

    This is free software: you can redistribute it and/or modify
    it under the terms of the GNU General Public License as published by
    the Free Software Foundation, either version 3 of the License, or
    (at your option) any later version.

    This software is distributed in the hope that it will be useful,
    but WITHOUT ANY WARRANTY; without even the implied warranty of
    MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE. See the
    GNU General Public License for more details.

    You should have received a copy of the GNU General Public License
    along with this software. If not, see <https://www.gnu.org/licenses/>.

===============================================================================

Solution class. Represents a potential solution. (The actual solving is done by
the Problem class.)

"""

import csv
import datetime
import logging
import os
from statistics import mean, median, variance
import sys
from typing import Dict, Generator, List, Tuple, TYPE_CHECKING

from cardinal_pythonlib.cmdline import cmdline_quote
from openpyxl.workbook.workbook import Workbook

from pdn_project_allocation.constants import (
    CsvHeadings,
    EXT_XLSX,
    SheetHeadings,
    SheetNames,
    SheetText,
)
from pdn_project_allocation.helperfunc import (
    autosize_openpyxl_column,
    autosize_openpyxl_worksheet_columns,
    bold_cell,
    bold_first_row,
)
from pdn_project_allocation.project import Project
from pdn_project_allocation.project_popularity import ProjectPopularity
from pdn_project_allocation.student import Student
from pdn_project_allocation.supervisor import Supervisor
from pdn_project_allocation.version import VERSION, VERSION_DATE

if TYPE_CHECKING:
    from pdn_project_allocation.problem import Problem

log = logging.getLogger(__name__)


# =============================================================================
# Solution
# =============================================================================


class Solution:
    """
    Represents a potential solution.
    """

    def __init__(
        self, problem: "Problem", allocation: Dict[Student, Project]
    ) -> None:
        """
        Args:
            problem:
                The :class:`Problem`, defining projects and students.
            allocation:
                The mapping of students to projects.
        """
        self.problem = problem
        self.allocation = allocation

    # -------------------------------------------------------------------------
    # Representations
    # -------------------------------------------------------------------------

    def __str__(self) -> str:
        """
        String representation.
        """
        lines = ["Solution:"]
        for student, project in self._gen_student_project_pairs():
            std = student.dissatisfaction(project)
            svd = project.dissatisfaction(student)
            lines.append(
                f"{student.shortname()} -> {project} "
                f"(student dissatisfaction {std}; "
                f"supervisor dissatisfaction {svd})"
            )
        return "\n".join(lines)

    def shortdesc(self) -> str:
        """
        Very short description. Ordered by student number.
        """
        students = sorted(self.allocation.keys(), key=lambda s: s.number)
        parts = [f"{s.number}: {self.allocation[s].number}" for s in students]
        return (
            "{"
            + ", ".join(parts)
            + "}"
            + ", student dissatisfaction "
            + f"{self.student_dissatisfaction_scores()}"
        )

    # -------------------------------------------------------------------------
    # Allocations
    # -------------------------------------------------------------------------

    def allocated_project(self, student: Student) -> Project:
        """
        Which project was allocated to this student?
        """
        return self.allocation[student]

    def allocated_students(self, project: Project) -> List[Student]:
        """
        Which students were allocated to this project?
        """
        return sorted(k for k, v in self.allocation.items() if v == project)

    def is_allocated(self, student: Student, project: Project) -> bool:
        """
        Is this student allocated to this project?
        """
        return self.allocation[student] == project

    def _gen_student_project_pairs(
        self,
    ) -> Generator[Tuple[Student, Project], None, None]:
        """
        Generates ``student, project`` pairs in student order.
        """
        students = sorted(self.allocation.keys(), key=lambda s: s.number)
        for student in students:
            project = self.allocation[student]
            yield student, project

    def n_students_allocated_to_project(self, project: Project) -> int:
        """
        How many students is this project allocated?
        """
        return len(self.allocated_students(project))

    def n_students_allocated_to_supervisor(
        self, supervisor: Supervisor
    ) -> int:
        """
        How many students is this supervisor allocated?
        """
        return sum(
            self.n_students_allocated_to_project(project)
            for project in self.problem.projects
            if project.is_supervised_by(supervisor)
        )

    def n_projects_allocated_to_supervisor(
        self, supervisor: Supervisor
    ) -> int:
        """
        How many projects is this supervisor allocated?
        """
        return sum(
            (1 if self.n_students_allocated_to_project(project) > 0 else 0)
            for project in self.problem.projects
            if project.is_supervised_by(supervisor)
        )

    # -------------------------------------------------------------------------
    # Student dissatisfaction
    # -------------------------------------------------------------------------

    def student_dissatisfaction_scores(self) -> List[float]:
        """
        All dissatisfaction scores.
        """
        dscores = []  # type: List[float]
        for student in self.problem.students:
            project = self.allocation[student]
            dscores.append(student.dissatisfaction(project))
        return dscores

    def student_dissatisfaction_median(self) -> float:
        """
        Median dissatisfaction per student.
        """
        return median(self.student_dissatisfaction_scores())

    def student_dissatisfaction_mean(self) -> float:
        """
        Mean dissatisfaction per student.
        """
        return mean(self.student_dissatisfaction_scores())

    def student_dissatisfaction_variance(self) -> float:
        """
        Variance of dissatisfaction scores.
        """
        return variance(self.student_dissatisfaction_scores())

    def student_dissatisfaction_min(self) -> float:
        """
        Minimum of dissatisfaction scores.
        """
        return min(self.student_dissatisfaction_scores())

    def student_dissatisfaction_max(self) -> float:
        """
        Maximum of dissatisfaction scores.
        """
        return max(self.student_dissatisfaction_scores())

    # -------------------------------------------------------------------------
    # Supervisor dissatisfaction
    # -------------------------------------------------------------------------

    def supervisor_dissatisfaction_scores_sum_students(self) -> List[float]:
        """
        All dissatisfaction scores. (If a project has several students, it
        scores the SUM of its dissatisfaction for each of those students
        scores.)
        """
        dscores = []  # type: List[float]
        for project in self.problem.projects:
            dscore = 0
            for student in self.problem.students:
                if self.allocation[student] == project:
                    dscore += project.dissatisfaction(student)
            dscores.append(dscore)
        return dscores

    def supervisor_dissatisfaction_scores_each_student(self) -> List[float]:
        """
        All dissatisfaction scores. (If a project has several students,
        multiple scores are returned for that project.)
        """
        dscores = []  # type: List[float]
        for project in self.problem.projects:
            for student in self.problem.students:
                if self.allocation[student] == project:
                    dscores.append(project.dissatisfaction(student))
        return dscores

    def supervisor_dissatisfaction_median(self) -> float:
        """
        Median dissatisfaction per student.
        """
        return median(self.supervisor_dissatisfaction_scores_each_student())

    def supervisor_dissatisfaction_mean(self) -> float:
        """
        Mean dissatisfaction per student.
        """
        return mean(self.supervisor_dissatisfaction_scores_each_student())

    def supervisor_dissatisfaction_variance(self) -> float:
        """
        Variance of dissatisfaction scores.
        """
        return variance(self.supervisor_dissatisfaction_scores_each_student())

    def supervisor_dissatisfaction_min(self) -> float:
        """
        Minimum of dissatisfaction scores.
        """
        return min(self.supervisor_dissatisfaction_scores_each_student())

    def supervisor_dissatisfaction_max(self) -> float:
        """
        Maximum of dissatisfaction scores.
        """
        return max(self.supervisor_dissatisfaction_scores_each_student())

    # -------------------------------------------------------------------------
    # Stability test
    # -------------------------------------------------------------------------

    def gen_better_projects(
        self, student: Student, project: Project
    ) -> Generator[Project, None, None]:
        """
        Generates projects that this student prefers over the specified one.
        """
        for p in self.problem.gen_better_projects(student, project):
            yield p

    def gen_better_students(
        self, project: Project, student: Student
    ) -> Generator[Student, None, None]:
        """
        Generates students that this project prefers over the specified one,
        for which they're eligible, AND who are are not already allocated to
        that project (bearing in mind that a project can have several
        students).
        """
        for s in self.problem.gen_better_students(project, student):
            if not self.is_allocated(s, project):
                yield s

    def stability(
        self, describe_all_failures: bool = True
    ) -> Tuple[bool, str]:
        """
        Is the solution a stable match, and if not, why not? See README.rst for
        discussion. See also https://gist.github.com/joyrexus/9967709.

        Arguments:
            describe_all_failures:
                Show all reasons for failure.

        Returns:
            tuple: (stable, reason_for_instability)
        """
        stable = True
        instability_reasons = []  # type: List[str]
        for student, project in self._gen_student_project_pairs():
            for alt_project in self.gen_better_projects(student, project):
                for alt_proj_student in self.allocated_students(alt_project):
                    if alt_proj_student == student:
                        continue
                    if student in self.gen_better_students(
                        alt_project, alt_proj_student
                    ):
                        instability_reasons.append(
                            f"Pairing of student {student} to project "
                            f"{project} is unstable. "
                            f"The student would rather have alternative "
                            f"project {alt_project}, and that alternative "
                            f"project would rather have {student} than their "
                            f"current allocation of {alt_proj_student}."
                        )
                        stable = False
                        if not describe_all_failures:
                            return False, "\n\n".join(instability_reasons)
        if stable:
            return True, "[Stable]"
        else:
            return False, "\n\n".join(instability_reasons)

    def is_stable(self) -> bool:
        """
        Is the solution a stable match?
        """
        return self.stability(describe_all_failures=False)[0]

    # -------------------------------------------------------------------------
    # Saving
    # -------------------------------------------------------------------------

    def write_xlsx(self, filename: str) -> None:
        """
        Writes the solution to an Excel XLSX file (and its problem, for data
        safety).

        Args:
            filename:
                Name of file to write.
        """
        log.info(f"Writing output to: {filename}")

        problem = self.problem

        # Not this way -- we can't then set column widths.
        #   wb = Workbook(write_only=True)  # doesn't create default sheet
        # Instead:
        wb = Workbook()
        wb.remove(wb.worksheets[0])

        # ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
        # Allocations, by student
        # ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
        ss = wb.create_sheet(SheetNames.STUDENT_ALLOCATIONS)
        n_students_assigned_project_not_preferred = 0
        n_students_assigned_supervisor_not_preferred = 0
        ss.append(
            [
                SheetHeadings.STUDENT,
                SheetHeadings.PROJECT,
                SheetHeadings.SUPERVISOR,
                SheetHeadings.STUDENT_PREFERENCE,
                SheetHeadings.NOT_PREFERRED_PROJECT,
                SheetHeadings.NOT_PREFERRED_SUPERVISOR,
            ]
        )
        for student, project in self._gen_student_project_pairs():
            if student.explicitly_ranked_project(project):
                _unhappy_project = SheetText.STUDENT_HAPPY
            else:
                _unhappy_project = SheetText.STUDENT_UNHAPPY_PROJECT
                n_students_assigned_project_not_preferred += 1
            if student.explicitly_ranked_any_supervisor(
                project.supervisors, problem.projects
            ):
                _unhappy_supervisor = SheetText.STUDENT_HAPPY
            else:
                _unhappy_supervisor = SheetText.STUDENT_UNHAPPY_SUPERVISOR
                n_students_assigned_supervisor_not_preferred += 1
            ss.append(
                [
                    student.name,
                    project.title,
                    project.supervisor_name(),
                    student.dissatisfaction(project),
                    _unhappy_project,
                    _unhappy_supervisor,
                ]
            )
        autosize_openpyxl_worksheet_columns(ss)
        bold_first_row(ss)

        # ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
        # Allocations, by project
        # ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
        ps = wb.create_sheet(SheetNames.PROJECT_ALLOCATIONS)
        ps.append(
            [
                SheetHeadings.PROJECT,
                SheetHeadings.SUPERVISOR,
                SheetHeadings.N_STUDENTS_ALLOCATED,
                SheetHeadings.STUDENTS,
                "Students' rank(s) of allocated project"
                " (dissatisfaction score)",
                "Project supervisor's rank(s) of allocated student(s)"
                " (dissatisfaction score)",
            ]
        )
        for project in problem.sorted_projects():
            student_names = []  # type: List[str]
            supervisor_dissatisfactions = []  # type: List[float]
            student_dissatisfactions = []  # type: List[float]
            for student in self.allocated_students(project):
                student_names.append(student.name)
                supervisor_dissatisfactions.append(
                    project.dissatisfaction(student)
                )
                student_dissatisfactions.append(
                    student.dissatisfaction(project)
                )
            ps.append(
                [
                    project.title,
                    project.supervisor_name(),
                    self.n_students_allocated_to_project(project),
                    ", ".join(student_names),
                    ", ".join(str(x) for x in student_dissatisfactions),
                    ", ".join(str(x) for x in supervisor_dissatisfactions),
                ]
            )
        autosize_openpyxl_worksheet_columns(ps)
        bold_first_row(ps)

        # ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
        # Allocations, by supervisor
        # ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~

        sa = wb.create_sheet(SheetNames.SUPERVISOR_ALLOCATIONS)
        sa.append(
            [
                SheetHeadings.SUPERVISOR,
                SheetHeadings.MAX_NUMBER_OF_PROJECTS,
                SheetHeadings.N_PROJECTS_ALLOCATED,
                SheetHeadings.MAX_NUMBER_OF_STUDENTS,
                SheetHeadings.N_STUDENTS_ALLOCATED,
            ]
        )
        for supervisor in problem.sorted_supervisors():
            sa.append(
                [
                    supervisor.name,
                    supervisor.max_n_projects,
                    self.n_projects_allocated_to_supervisor(supervisor),
                    supervisor.max_n_students,
                    self.n_students_allocated_to_supervisor(supervisor),
                ]
            )
        autosize_openpyxl_worksheet_columns(sa)
        bold_first_row(sa)

        # ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
        # Popularity of projects
        # ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
        pp = wb.create_sheet(SheetNames.PROJECT_POPULARITY)
        pp.append(ProjectPopularity.headings())
        popularities = []  # type: List[ProjectPopularity]
        for project in problem.projects:
            popularities.append(ProjectPopularity(project, self))
        ProjectPopularity.sort_and_assign_ranks(popularities)
        for projpop in popularities:
            pp.append(projpop.values())
        autosize_openpyxl_column(pp, 1)  # the supervisor column
        autosize_openpyxl_column(pp, 6)  # the student column
        # Not all columns (lengthy project titles at left).
        bold_first_row(pp)

        # ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
        # Applied but ineligible -- suggests communication failure
        # ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
        ai = wb.create_sheet(SheetNames.APPLIED_BUT_INELIGIBLE)
        ai.append(
            [
                SheetHeadings.STUDENT,
                SheetHeadings.STUDENT_PREFERENCE,
                SheetHeadings.ELIGIBLE,
                SheetHeadings.SUPERVISOR,
                SheetHeadings.PROJECT,
            ]
        )
        # Project last, then we can autosize and see everything.
        for project in problem.sorted_projects():
            for student in problem.sorted_students():
                if student.explicitly_ranked_project(
                    project
                ) and not problem.eligibility.is_eligible(student, project):
                    ai.append(
                        [
                            student.name,
                            student.dissatisfaction(project),
                            0,
                            project.supervisor_name(),
                            project.title,
                        ]
                    )
        autosize_openpyxl_worksheet_columns(ai)
        bold_first_row(ai)

        # ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
        # Software, settings, and summary information
        # ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
        zs = wb.create_sheet(SheetNames.INFORMATION)
        is_stable, instability_reason = self.stability()
        zs_rows = [
            ["SOFTWARE DETAILS"],
            [],
            ["Software", "pdn_project_allocation"],
            ["Version", VERSION],
            ["Version date", VERSION_DATE],
            [
                "Source code",
                "https://github.com/RudolfCardinal/pdn_project_allocation",
            ],
            ["Author", "Rudolf Cardinal (rudolf@pobox.com)"],
            [],
            ["RUN INFORMATION"],
            [],
            ["Date/time", datetime.datetime.now()],
            [
                "Overall weight given to student preferences",
                1 - self.problem.config.supervisor_weight,
            ],
            [
                "Overall weight given to supervisor preferences",
                self.problem.config.supervisor_weight,
            ],
            ["Command-line parameters", cmdline_quote(sys.argv)],
            ["Config", str(self.problem.config)],
            [],
            ["SUMMARY STATISTICS"],
            [],
            [
                "Student dissatisfaction median",
                self.student_dissatisfaction_median(),
            ],
            [
                "Student dissatisfaction mean",
                self.student_dissatisfaction_mean(),
            ],
            [
                "Student dissatisfaction variance",
                self.student_dissatisfaction_variance(),
            ],
            [
                "Student dissatisfaction minimum",
                self.student_dissatisfaction_min(),
            ],
            [
                "Student dissatisfaction minimum",
                self.student_dissatisfaction_max(),
            ],
            [
                "Number of students not assigned a preferred project",
                n_students_assigned_project_not_preferred,
            ],
            [
                "Number of students not assigned a preferred supervisor",
                n_students_assigned_supervisor_not_preferred,
            ],
            [],
            [
                "Supervisor dissatisfaction (with each student) median",
                self.supervisor_dissatisfaction_median(),
            ],
            [
                "Supervisor dissatisfaction (with each student) mean",
                self.supervisor_dissatisfaction_mean(),
            ],
            [
                "Supervisor dissatisfaction (with each student) variance",
                self.supervisor_dissatisfaction_variance(),
            ],
            [
                "Supervisor dissatisfaction (with each student) minimum",
                self.supervisor_dissatisfaction_min(),
            ],
            [
                "Supervisor dissatisfaction (with each student) maximum",
                self.supervisor_dissatisfaction_max(),
            ],
            [],
            ["Stable marriages?", str(is_stable)],
            ["If unstable, reason:", instability_reason],
        ]
        for row in zs_rows:
            zs.append(row)
        autosize_openpyxl_column(zs, 0)
        # Not all columns: right-justification would make numbers invisible in
        # the presence of some long text. But we can make the date/time
        # plainer:
        zs.column_dimensions["B"].width = 20
        bold_first_row(zs)
        bold_cell(zs["A9"])
        bold_cell(zs["A17"])

        # ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
        # Problem definition
        # ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
        self.problem.write_to_xlsx_workbook(wb, with_internal_prefs=True)

        wb.save(filename)
        wb.close()

    def write_data(self, filename: str) -> None:
        """
        Autodetects the file type from the extension and writes data to that
        file.
        """
        # File type?
        _, ext = os.path.splitext(filename)
        if ext == EXT_XLSX:
            self.write_xlsx(filename)
        else:
            raise ValueError(
                f"Don't know how to write file type {ext!r} for {filename!r}"
            )

    def write_student_csv(self, filename: str) -> None:
        """
        Writes just the "per student" mapping to a CSV file, for comparisons
        (e.g. via ``meld``).
        """
        log.info(f"Writing student allocation data to: {filename}")
        with open(filename, "w") as file:
            writer = csv.writer(file)
            writer.writerow(
                [
                    CsvHeadings.STUDENT_NUMBER,
                    CsvHeadings.STUDENT_NAME,
                    CsvHeadings.PROJECT_NUMBER,
                    CsvHeadings.PROJECT_NAME,
                    CsvHeadings.DISSATISFACTION_SCORE,
                ]
            )
            for student, project in self._gen_student_project_pairs():
                writer.writerow(
                    [
                        student.number,
                        student.name,
                        project.number,
                        project.title,
                        student.dissatisfaction(project),
                    ]
                )
